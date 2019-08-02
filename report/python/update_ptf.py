# -*- coding: utf-8 -*-
"""
Created on Fri Jan  4 12:56:14 2019

@author: Giulio Rossetti
"""
import sys
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import FuncFormatter
import datetime as dt
import seaborn as sns
import xlwings as xw
from openpyxl import Workbook #pip install openpyxl
from openpyxl import load_workbook
import matplotlib.ticker as mtick
from pandas.tseries.offsets import BDay
from matplotlib.ticker import StrMethodFormatter
import arch
import operator
import os
#file =r"Y:\Mobiliare\08 Finint Economia Reale Italia\01_Front Office\02 Gestione\Scripts Python\PTF PIR.csv"
#saveLoc = "Y:\\Mobiliare\\08 Finint Economia Reale Italia\\00_Documenti_Reportistica\\03 Comitati\\Immagini update portafoglio\\"
#ext = ".pdf"


#plt.style.use('ggplot')
#plt.style.use('classic')
to_report = True
rf=0.00
dpi_report=150
def equidate_ax(fig, ax, dates, fmt="%Y-%m-%d", label="Date"):
    """
    Sets all relevant parameters for an equidistant date-x-axis.
    Tick Locators are not affected (set automatically)

    Args:
        fig: pyplot.figure instance
        ax: pyplot.axis instance (target axis)
        dates: iterable of datetime.date or datetime.datetime instances
        fmt: Display format of dates
        label: x-axis label
    Returns:
        None

    """    
    N = len(dates)
    def format_date(index, pos):
        index = np.clip(int(index + 0.5), 0, N - 1)
        return dates[index].strftime(fmt)
    ax.xaxis.set_major_formatter(FuncFormatter(format_date))
    ax.set_xlabel(label)
    fig.autofmt_xdate()

def double_pivot(data,value,key):
    piv = pd.pivot_table(data,value,key,aggfunc=np.sum)
    df = pd.concat([piv,piv.divide(piv.sum())],axis=1,sort=True)
    df.columns = [value,'%']
    return df

def beauty_pie_4(d00,d01,d10,d11,title):
    def set_title_beauty_pie(tit):
        ax1.set_title(tit,fontdict =  {'family': 'serif',
        'color':  'darkblue',
        'weight': 'normal',
        'size': 35,
        },pad=60,loc='center') 
        return 
    
    fig = plt.figure(figsize=(40,27), dpi=dpi_report)
    colors = ['darkblue','steelblue','royalblue','midnightblue']
    size =0.4

    labels1 = [r.replace(' ','-\n') if len(r) > 15 else r for r in d00.index.tolist()]
    labels2 = [r.replace(' ','-\n') if len(r) > 15 else r for r in d01.index.tolist()]
    labels3 = [r.replace(' ','-\n') if len(r) > 15 else r for r in d10.index.tolist()]
    labels4 = [r.replace(' ','-\n') if len(r) > 15 else r for r in d11.index.tolist()]
    
    size0=25
    size1 = 20
    
    ax1 = plt.subplot2grid((2,2),(0,0))
    set_title_beauty_pie(title[0]) 
    patches, texts, autotexts = ax1.pie(d00['MKT VALUE'],autopct='%1.1f%%',pctdistance=0.8, wedgeprops=dict(width=size, edgecolor='w'),colors=colors,labels =labels1,)
    plt.setp(autotexts, size=size0, weight="bold",color='w')
    plt.setp(texts, size=size1, weight="bold", color='navy')
    ax1.set(aspect="equal")
    
    ax1 = plt.subplot2grid((2,2),(0,1))
    set_title_beauty_pie(title[1]) 
    patches, texts, autotexts = ax1.pie(d01['MKT VALUE'],autopct='%1.1f%%',pctdistance=0.8, wedgeprops=dict(width=size, edgecolor='w'),colors=colors,labels =labels2,)
    plt.setp(autotexts, size=size0, weight="bold",color='w')
    plt.setp(texts, size=size1, weight="bold", color='navy')
    ax1.set(aspect="equal")
    ax1.set(aspect="equal")    
    
    ax1 = plt.subplot2grid((2,2),(1,0))
    set_title_beauty_pie(title[2]) 
    patches, texts, autotexts = ax1.pie(d10['MKT VALUE'],autopct='%1.1f%%',pctdistance=0.8, wedgeprops=dict(width=size, edgecolor='w'),colors=colors,labels =labels3,)
    plt.setp(autotexts, size=size0, weight="bold",color='w')
    plt.setp(texts, size=size1, weight="bold", color='navy')
    ax1.set(aspect="equal")    
    
    ax1 = plt.subplot2grid((2,2),(1,1))
    set_title_beauty_pie(title[3]) 
    patches, texts, autotexts = ax1.pie(d11['MKT VALUE'],autopct='%1.1f%%',pctdistance=0.8, wedgeprops=dict(width=size, edgecolor='w'),colors=colors,labels =labels4,)
    plt.setp(autotexts, size=size0, weight="bold",color='w')
    plt.setp(texts, size=size1, weight="bold", color='navy')
    ax1.set(aspect="equal")    
    
    plt.tight_layout()  
    plt.close(fig)
    return fig

def maxdrawdown(performance):
    md = []
    for i in range(np.size(performance,1)):
        trough = np.argmax(np.maximum.accumulate(performance.iloc[:,i]) - performance.iloc[:,i])
        peak = np.argmax(performance.iloc[:,i][:trough])
    
        md_a = performance.iloc[:,i][trough]/performance.iloc[:,i][peak]-1
        md.append(md_a)
    return md

def beauty_pie_2(d00,d10,title):
    def set_title_beauty_pie(tit):
        ax1.set_title(tit,fontdict =  {'family': 'serif',
        'color':  'darkblue',
        'weight': 'normal',
        'size': 35,
        },pad=60,loc='center') 
        return 
    
    fig = plt.figure(figsize=(20,27), dpi=dpi_report)
    colors = ['darkblue','steelblue','royalblue','midnightblue']
    size =0.4

    labels1 = [r.replace(' ','-\n') if len(r) > 15 else r for r in d00.index.tolist()]
    labels2 = [r.replace(' ','-\n') if len(r) > 15 else r for r in d10.index.tolist()]
    
    size0=25
    size1 = 20
    
    ax1 = plt.subplot2grid((2,1),(0,0))
    set_title_beauty_pie(title[0]) 
    patches, texts, autotexts = ax1.pie(d00['MKT VALUE'],autopct='%1.1f%%',pctdistance=0.8, wedgeprops=dict(width=size, edgecolor='w'),colors=colors,labels =labels1,)
    plt.setp(autotexts, size=size0, weight="bold",color='w')
    plt.setp(texts, size=size1, weight="bold", color='navy')
    ax1.set(aspect="equal")
       
    ax1 = plt.subplot2grid((2,1),(1,0))
    set_title_beauty_pie(title[1]) 
    patches, texts, autotexts = ax1.pie(d10['MKT VALUE'],autopct='%1.1f%%',pctdistance=0.8, wedgeprops=dict(width=size, edgecolor='w'),colors=colors,labels =labels2,)
    plt.setp(autotexts, size=size0, weight="bold",color='w')
    plt.setp(texts, size=size1, weight="bold", color='navy')
    ax1.set(aspect="equal")    
    
    plt.tight_layout()  
    plt.close(fig)
    return fig


def beuty_pie(df,title):
    fig,ax = plt.subplots(num=None, figsize=(16, 13) ,dpi=dpi_report)
    colors = ['darkblue','steelblue','royalblue','midnightblue']
    size = 0.4
    labels = [r.replace(' ','-\n') if len(r) > 15 else r for r in df.index.tolist()]
    patches, texts, autotexts = ax.pie(df['MKT VALUE'],autopct='%1.1f%%',radius=1.25, wedgeprops=dict(width=size, edgecolor='w'),colors=colors,labels =labels )
#    ax.set(aspect="equal",anchor='C')
    ax.set_title(title,fontdict =  {'family': 'serif',
        'color':  'darkblue',
        'weight': 'normal',
        'size': 30,
        },pad=60,loc='center') 

    return fig,title
        
def cond(x):
    if x <1:
        return '<1year'
    elif x > 1 and x < 5:
        return '1-5 years'
    elif x > 5:
        return '>5 years'
    elif np.isnan(x):
        return 'perp/other'

def cond1(x):
    if x <1:
        return '<1year'
    elif x > 1 and x < 5:
        return '1-5 years'
    elif x > 5:
        return '>5 years'
    elif np.isnan(x):
        return 'n.d.'
    
def cond_div(x):
    if x <0.01:
        return '<0.1%'
    elif x > 0.01 and x < 0.03:
        return '1%-3%'
    elif x > 0.03 and x<0.05:
        return '3%-5%'
    elif x > 0.05:
        return '>5%'
    elif np.isnan(x):
        return 'ETF or n.d.'

def cond_cap(x):
    if x <500e06:
        return '<100 Mln'
    elif x > 500e06 and x < 1000e06:
        return '500-1000 Mln'
    elif x > 1000e06:
        return '>1000 Mln'
    else:
        return 'ETF and Others'
    
    
if to_report:
    name_report = dt.datetime.now().strftime("%d.%m.%Y")
    year_save = dt.datetime.now().strftime("%Y")
    month_save = dt.datetime.now().strftime("%m.%y")
    name_folder = 'Y:\\Mobiliare\\08 Finint Economia Reale Italia\\01_Front Office\\02 Gestione\\Report\\' + year_save + '\\'+month_save + '\\'
    if not os.path.exists(name_folder):
        os.makedirs(name_folder)
    
    
    name_to_save = name_folder + 'FERI_REPORT_'+name_report+'.xlsm'

    wb1 = Workbook()
    wb1= load_workbook(filename = r'Y:\Mobiliare\08 Finint Economia Reale Italia\01_Front Office\02 Gestione\Report\template.xlsm',read_only=False,keep_vba=True)
    wb1.save(name_to_save) 
    
    wb = Workbook()
    wb= load_workbook(filename = name_to_save,read_only=False,keep_vba=True)


    writer = pd.ExcelWriter(name_to_save, engine='openpyxl')
    writer.book = wb
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

    
old_stdout = sys.stdout
savefig = False
s_flag = input('save images? (y/n) ')
to_text = input('save to txt? (y/n) ')

if to_text == "y":
    sys.stdout = open("Y:\\Mobiliare\\08 Finint Economia Reale Italia\\00_Documenti_Reportistica\\03 Comitati\\Immagini update portafoglio\\results_"+dt.datetime.now().strftime("%d_ %m_%y")+".txt", "w")
#else:
#    sys.stdout = sys.__stdout__
#
if s_flag == "y":
    savefig=True
sys.__stdout__
str_pir_perf = "PIR_PERF.csv"
str_bond_val ="BOND.csv"
str_ptf_data = 'PTF PIR.csv'
str_eqt_price = 'EQUITY_price.CSV'
str_eqt_pos = 'EQUITY_pos.CSV'
str_divfile = 'dividends.xlsx'

#reading data from csv and dropping cells containing ""
pir_perf = pd.read_csv(str_pir_perf,encoding = "ISO-8859-1").dropna(how='all')
bond_val = pd.read_csv(str_bond_val ,encoding = "ISO-8859-1").dropna(how='all')
ptf_data = pd.read_csv(str_ptf_data ,encoding = "ISO-8859-1").dropna(how='all')
eqt_price = pd.read_csv(str_eqt_price ,encoding = "ISO-8859-1").dropna(how='all')
eqt_pos = pd.read_csv(str_eqt_pos ,encoding = "ISO-8859-1").dropna(how='all')

div_read = pd.ExcelFile(str_divfile)
div_data = div_read.parse('dividends')
#set Date as index and convert it to integer
pir_perf.set_index('Date',inplace=True) 
bond_val.set_index('Date',inplace=True) 
ptf_data.set_index('Date',inplace=True) 
eqt_price.set_index('Date',inplace=True) 
eqt_pos.set_index('Date',inplace=True)

pir_perf.index=pir_perf.index.map(int)
bond_val.index = bond_val.index.map(int)
ptf_data.index = ptf_data.index.map(int)
eqt_price.index = eqt_price.index.map(int)
eqt_pos.index = eqt_pos.index.map(int)

#check for labels and indeces inconsistency

print(dt.datetime.now().strftime("%A, %d %B %Y %H:%M:%S"),"\n")  
print("="*60)    
print(" "*23+"START PROGRAM:")
print("="*60) 
print("As of ", pir_perf.index[-1])

if eqt_price.columns.all() == eqt_pos.columns.all():
    print ("\nMatching labels for equity prices and positions: proceed!")
else:
    sys.exit("ERROR: Equity labels don't match")
        
if eqt_price.index.all() == eqt_pos.index.all():
    print ("Matching indeces for equity prices and positions: proceed!\n")
else:
    sys.exit("ERROR: Eqiuity indeces don't match")   

  
print("="*60)    
print("Checking if dataframe indeces are consistent:")
print("="*60) 

print("Equity positions: Length index = ",np.size(eqt_pos,0))
print("Equity prices: Length index = ",np.size(eqt_pos,0))
print("Bond market values: Length index = ",np.size(bond_val,0))
print("PIR Fund: Length index = ",np.size(pir_perf,0))

if eqt_price.index.all() == bond_val.index.all()==pir_perf.index.all():
    print ("\nMatching indeces for equity, bond and PIR fund: proceed!\n")
else:
    sys.exit("ERROR: indeces don't match")
    
#extract benchmark prices from the PTD PIR csv
benchmark = ['FTSEMIB Index','ITPIRMC Index','ITLMS Index','FITSC Index','ITMC Index','PIRB IM Equity']
berk_price = ptf_data[ptf_data.columns.intersection(benchmark)]
berk_price = berk_price.loc[berk_price.index.intersection(eqt_price.index)]

#compute equity and benchmark returns
eqt_ret_no_adj = (eqt_price/eqt_price.shift(1) - 1).dropna(0,how='all')
berk_ret = (berk_price/berk_price.shift(1) - 1).dropna(0,how='all')

#compute historical market value as price * quantity and asset class cap
eqt_val = (eqt_price * eqt_pos).fillna(0)
eqt_cap = np.sum(eqt_val,axis = 1).rename('eqt_cap')
eqt_w = eqt_val.divide(eqt_cap,0).fillna(0)
bond_cap = np.sum(bond_val,axis = 1).rename('bond_cap')
fund_cap = eqt_cap + bond_cap
w_asset_class = pd.concat([eqt_cap,bond_cap],axis=1).divide(fund_cap,0)
#adjusting ret for dividends
div_data['Ex-Date'] = div_data['Ex-Date'].dt.strftime('%Y%m%d')
div_data['Ex-Date'] =div_data['Ex-Date'].astype(int)

for i in range(np.size(div_data,0)):
    try:
        idx_div = eqt_price.index.get_loc(div_data.loc[i,'Ex-Date'])-1
        if div_data.loc[i,'Dividend Type'] =="Regular Cash" or div_data.loc[i,'Dividend Type'] =="Rights Issue":
            div_data.loc[i,'price_t-1'] =eqt_price.loc[eqt_price.index[idx_div],div_data.loc[i,'Security']]
        else:
            div_data.loc[i,'price_t-1'] = np.nan
        
    except:
        pass
    
div_data['div_yield'] = div_data['Dividend Amount']/div_data['price_t-1']
div_ret = pd.DataFrame().reindex_like(eqt_ret_no_adj)
div_gross = pd.DataFrame().reindex_like(eqt_ret_no_adj)

for j in range(np.size(div_data,0)):
    try:
        div_ret.loc[div_data.loc[j,'Ex-Date'],div_data.loc[j,'Security']] = div_data.loc[j,'div_yield']
        div_gross.loc[div_data.loc[j,'Ex-Date'],div_data.loc[j,'Security']] = div_data.loc[j,'div_yield'] * div_data.loc[j,'price_t-1']
    except:
        pass
        
div_ret = div_ret.fillna(0)
div_ret = div_ret.loc[eqt_ret_no_adj.index,:]

div_gross = div_gross.fillna(0)
div_gross = div_gross.loc[eqt_ret_no_adj.index,:]

div_amount = div_gross * eqt_pos
div_amount = div_amount.fillna(0)
div_2017 = div_amount[ (div_amount.index<20180101)].sum()
div_2018 = div_amount[(div_amount.index>20171231) & (div_amount.index<20190101)].sum()

eqt_ret = eqt_ret_no_adj + div_ret


#compute portfolio equity retuns
eqt_ptf_ret = []
for i in range(len(eqt_ret)):
    r_e = eqt_w.iloc[i,:].T @ eqt_ret.iloc[i,:].fillna(0)
    eqt_ptf_ret.append(r_e)

eqt_ptf_ret = pd.DataFrame(eqt_ptf_ret,index=eqt_ret.index,columns=['port_eqt_ret'])

#computing bond performance from equity perf, PIR perf and weights
#starting from PIR PERF 20170707 and eqt w and ret 20170706
bond_ptf_ret = []
for i in range(len(eqt_ret)-1):
    r_b = ((pir_perf.iloc[i+2,1] - w_asset_class.iloc[i+1,0] * eqt_ptf_ret.iloc[i+1,0])\
        / w_asset_class.iloc[i+1,1]) + 0.00007352
    bond_ptf_ret.append(r_b)
bond_ptf_ret = pd.DataFrame(bond_ptf_ret,index=eqt_ret.index[1:],columns=['port_bond_ret'])

#normalizing index, first datapoint is 07/07/2017
eqt_ptf_ret = eqt_ptf_ret.loc[eqt_ptf_ret.index.intersection(bond_ptf_ret.index)]
berk_ret = berk_ret.loc[berk_ret.index.intersection(bond_ptf_ret.index)]
    
#computing cumulative performance for equity and bond portfolios and benchmarks
eqt_perf = eqt_ptf_ret.add(1).cumprod()
bond_perf = bond_ptf_ret.add(1).cumprod()
berk_perf = berk_ret.add(1).cumprod()
pir_perf_a = pd.DataFrame(pir_perf["Return"].add(1).cumprod())

df_perf_all = pd.concat([eqt_perf,bond_perf,pir_perf_a],sort=False,axis=1).fillna(1)
md = maxdrawdown(df_perf_all)

datedt = [20170706]
eqt_monthly = [[20170706,1]]
bond_monthly = [[20170706,1]]
pir_monthly = [[20170706,1]]
     
for i in range(np.size(eqt_perf,0)):
        if not str(eqt_perf.index[i])[4:6] == str(eqt_perf.index[i-1])[4:6] and i!=0:
            last_m = eqt_perf.index[i-1]
            datedt.append(last_m)
for i in range(len(datedt)-1):
    eqt_monthly.append([datedt[i+1],eqt_perf.loc[int(datedt[i+1])][0]])
    bond_monthly.append([datedt[i+1],bond_perf.loc[int(datedt[i+1])][0]])
    pir_monthly.append([datedt[i+1],pir_perf_a.loc[int(datedt[i+1])][0]])
                   
#mtd
eqt_mtd = (eqt_perf.iloc[-1]/eqt_perf.loc[int(last_m)]-1)[0]
bond_mtd = (bond_perf.iloc[-1]/bond_perf.loc[int(last_m)]-1)[0]
pir_mtd = (pir_perf_a.iloc[-1]/pir_perf_a.loc[int(last_m)]-1)[0]
mtd = [eqt_mtd,bond_mtd,pir_mtd]
mtd = pd.DataFrame(mtd,index=['Equity','Bond','PIR'],columns = ['MTD'])
#ytd
eqt_ytd = (eqt_perf.iloc[-1]/eqt_perf.loc[20181228]-1)[0]
bond_ytd = (bond_perf.iloc[-1]/bond_perf.loc[20181228]-1)[0]
pir_ytd = (pir_perf_a.iloc[-1]/pir_perf_a.loc[20181228]-1)[0]
ytd = [eqt_ytd,bond_ytd,pir_ytd]
ytd = pd.DataFrame(ytd,index=['Equity','Bond','PIR'],columns = ['YTD'])
#2018
eqt_2018 = (eqt_perf.loc[20181228]/eqt_perf.loc[20171229]-1)[0]
bond_2018 = (bond_perf.loc[20181228]/bond_perf.loc[20171229]-1)[0]
pir_2018 = (pir_perf_a.loc[20181228]/pir_perf_a.loc[20171229]-1)[0]
p_2018 = [eqt_2018,bond_2018,pir_2018]
p_2018 = pd.DataFrame(p_2018,index=['Equity','Bond','PIR'],columns = ['2018'])
#2017
eqt_2017 = (eqt_perf.loc[20171229]/1-1)[0]
bond_2017 =( bond_perf.loc[20171229]/1-1)[0]
pir_2017 = (pir_perf_a.loc[20171229]/1-1)[0]
p_2017 = [eqt_2017,bond_2017,pir_2017]
p_2017 = pd.DataFrame(p_2017,index=['Equity','Bond','PIR'],columns = ['2017'])

perf_summary =pd.concat([mtd,ytd,p_2018,p_2017],sort=False,axis=1,join='outer')*100

#constructing the dataframes with timestamp indeces         
eqt_monthly = pd.DataFrame(eqt_monthly,columns=['Date','Equity Monthly Ret']) 
eqt_monthly.set_index('Date',inplace=True)

bond_monthly = pd.DataFrame(bond_monthly,columns=['Date','Bond Monthly Ret']) 
bond_monthly.set_index('Date',inplace=True)

pir_monthly = pd.DataFrame(pir_monthly,columns=['Date','PIR Monthly Ret']) 
pir_monthly.set_index('Date',inplace=True)

timestamp_index = [dt.datetime.strptime(str(date),'%Y%m%d') for date in eqt_monthly.index[1:]]  
timestamp_index =[date.strftime("%b %Y") for date in timestamp_index]
#timestamp_index = timestamp_index[:-1]  
#
eqt_monthly = (eqt_monthly/eqt_monthly.shift(1)-1).dropna(how='any')
eqt_monthly.set_index([timestamp_index],inplace=True)
bond_monthly = (bond_monthly/bond_monthly.shift(1)-1).dropna(how='any')
bond_monthly.set_index([timestamp_index],inplace=True)
pir_monthly = (pir_monthly/pir_monthly.shift(1)-1).dropna(how='any')
pir_monthly.set_index([timestamp_index],inplace=True)

df_monthly_perf = pd.concat([eqt_monthly,bond_monthly,pir_monthly],axis=1)
df_cum_perf = df_monthly_perf.add(1).cumprod()*100
#printing monthly performance
#for i in range(len(df_monthly_perf)):
#    print('{:2}   {:2}%   {:2}% '.format(df_monthly_perf.index[i],round(df_monthly_perf.iloc[i,0]*100,2),\
#                  round(df_monthly_perf.iloc[i,1]*100,2)))
print("="*60)    
print("Printing monthly equity and bond returns:")
print("="*60) 
print("-"*40)
print('%s %12.6s %8.6s %8.6s'%('Date','Equity','Bond','PIR'))
print("-"*40) 
for i in range(len(df_monthly_perf)):
    print('%s %8.4f%% %8.4f%% %8.4f%% ' %(df_monthly_perf.index[i],df_monthly_perf.iloc[i,0]*100,\
                                  df_monthly_perf.iloc[i,1]*100,df_monthly_perf.iloc[i,2]*100))
#printing mtd ytd 2018 2017 perf
print("")
print("-"*40)

print(perf_summary.to_string(col_space=7,justify='center',line_width=12,formatters={'MTD':'{:,.3f}%'.format,
                             'YTD':'{:,.3f}%'.format,
                             '2018':'{:,.3f}%'.format,
                             '2017':'{:,.3f}%'.format}))
print("-"*40)

#printing monthly performance of €100 invested
print("="*60)    
print("Printing performance of €100 invested:")
print("="*60) 
print("-"*40)
print('%s %12.6s %8.6s %8.6s'%('Date','Equity','Bond','PIR'))
print("-"*40) 
for i in range(len(df_monthly_perf)):
    print('%s %9.3f %9.3f %9.3f' %(df_cum_perf.index[i],df_cum_perf.iloc[i,0],\
                                  df_cum_perf.iloc[i,1],df_cum_perf.iloc[i,2]))
print("\n")    
print("="*60)    
print("EQUITY: Industry Analysis")
print("="*60) 
print("") 

#TO DO        
"""

UPDATE QUARTERS

"""
#Equity Portfolio, industry analysis
eqt_ind = (ptf_data.loc[:,['TICKER','Industry Breakdown']]).dropna(how='all')\
            .fillna("ETF").reset_index().drop('Date',axis=1)
#a = set(eqt_ind['TICKER'].tolist())
#b = set(eqt_price.columns.tolist())

###TO DO: Add BNS EIT SRI to ptf_data
df_missing = pd.DataFrame([['BNS IM Equity','Financial'],[ 'EIT IM Equity','Communications'],\
                           ['SRI IM Equity','Consumer, Non-cyclical'],['ITD IM Equity','Communications'],['EXSY IM Equity','Technology']],columns=['TICKER','Industry Breakdown'])
eqt_ind = eqt_ind.append(df_missing,ignore_index=True)
eqt_ind.set_index('TICKER',inplace=True)
eqt_ind =eqt_ind.T
eqt_ind = eqt_ind[eqt_price.columns]
industry_names = pd.unique(eqt_ind.loc['Industry Breakdown',:]).tolist()
eqt_ind = eqt_ind.T
eqt_ind = eqt_ind.reset_index()

industry_weights_ts = []
n_col_industry = 0
for industry in industry_names:
    print(industry.upper())
    #retriving tickers belonging to selected industry
    ticker_s = eqt_ind[eqt_ind['Industry Breakdown']==industry]
 
    ind_eqt_val =  eqt_val[eqt_val.columns.intersection(ticker_s['TICKER'].tolist())]
    ind_eqt_ret = eqt_ret[eqt_ret.columns.intersection(ticker_s['TICKER'].tolist())]
    ind_eqt_cap = ind_eqt_val.sum(axis=1)
    ind_eqt_w = ind_eqt_val.divide(ind_eqt_cap,axis=0)
    
    
    w_ind = ind_eqt_cap.divide(eqt_cap,axis=0).fillna(0)
    w_ind.rename('Weight ' +industry ,inplace=True)
    industry_weights_ts.append(w_ind)
    #computing industry returns
    ind_eqt_ptf_ret = []
    for i in range(len(ind_eqt_ret)):
        r_ind = ind_eqt_w.iloc[i,:].fillna(0) @ ind_eqt_ret.iloc[i,:].fillna(0).T
        ind_eqt_ptf_ret.append(r_ind)
    ind_eqt_ptf_ret = pd.DataFrame(ind_eqt_ptf_ret,index=ind_eqt_ret.index,columns=['Ret ' + industry])
    ind_eqt_ptf_ret = ind_eqt_ptf_ret.loc[ind_eqt_ptf_ret.index.intersection(eqt_ptf_ret.index)]   
    #cumulative perf
    ind_eqt_perf = ind_eqt_ptf_ret.add(1).cumprod()
    #monthly perf
    ind_eqt_monthly = [[20170706,1]]
#    for i in range(np.size(ind_eqt_perf,0)):
#        if i ==0:
#            ind_eqt_monthly.append([ind_eqt_perf.index[i],1])
#        else:
#            if not str(ind_eqt_perf.index[i])[4:6] == str(ind_eqt_perf.index[i-1])[4:6]:
#                ind_eqt_monthly.append([ind_eqt_perf.index[i+1],ind_eqt_perf.iloc[i+1,0]])
    for i in range(len(datedt)-1):
        ind_eqt_monthly.append([datedt[i+1],ind_eqt_perf.loc[int(datedt[i+1])][0]])
     
    ind_eqt_monthly = pd.DataFrame(ind_eqt_monthly,columns=['Date','Monthly Ret' + industry]) 
    ind_eqt_monthly.set_index('Date',inplace=True)
   
    ind_eqt_monthly = (ind_eqt_monthly/ind_eqt_monthly.shift(1)-1).dropna(how='any')
    ind_eqt_monthly.set_index([timestamp_index],inplace=True)
    ind_eqt_mtd = (ind_eqt_perf.iloc[-1]/ind_eqt_perf.loc[int(last_m)]-1)[0]
    
    print("-"*40)
    print('%s %15.11s '%('Date',"Montlhy Ret"))
    print("-"*40) 
    for i in range(len(ind_eqt_monthly)):
        print('%s %8.4f%% ' %(ind_eqt_monthly.index[i],ind_eqt_monthly.iloc[i,0]*100))
    
    print("")
    print('N stocks = %15.i ' %(np.size(ind_eqt_ret,1)))
    print('Weight(EQUITY) = %8.4f%% ' %(ind_eqt_cap.iloc[-1]/eqt_cap.iloc[-1]*100))
    print('MTD = %19.4f%% ' %(ind_eqt_mtd*100))
    print("")
    print("END ", industry)    
    print("_"*40)
    print("")
    
    if to_report:
        ind_eqt_perf
        if n_col_industry==0:
            ind_eqt_perf.to_excel(writer,'Equity performance',startrow = 0,startcol = n_col_industry+5,index=True)
            w_ind.to_excel(writer,'Weights',startrow = 0,startcol = n_col_industry+5,index=True)
            n_col_industry +=2
        else:
            ind_eqt_perf.to_excel(writer,'Equity performance',startrow = 0,startcol = n_col_industry+5,index=False)
            w_ind.to_excel(writer,'Weights',startrow = 0,startcol =n_col_industry+5,index=False)
            n_col_industry +=1
            
    
industry_weights_ts = pd.concat(industry_weights_ts,axis=1)
industry_weights_ts = industry_weights_ts.loc[industry_weights_ts.index.intersection(eqt_ptf_ret.index)]
timestamp_index_long = [dt.datetime.strptime(str(date),'%Y%m%d') for date in industry_weights_ts.index]  
industry_weights_ts.set_index([timestamp_index_long],inplace=True)

SMALL_SIZE = 8
MEDIUM_SIZE = 10
BIGGER_SIZE = 14



#==================PLOTTING PARAMS================================
font = {'family': 'serif',
        'color':  'darkblue',
        'weight': 'normal',
        'size': 30,
        }

#plt.rc('font', size=SMALL_SIZE)          # controls default text sizes
#plt.rc('axes', titlesize=SMALL_SIZE)     # fontsize of the axes title
#plt.rc('axes', labelsize=SMALL_SIZE)    # fontsize of the x and y labels
#plt.rc('xtick', labelsize=MEDIUM_SIZE)    # fontsize of the tick labels
#plt.rc('ytick', labelsize=SMALL_SIZE)    # fontsize of the tick labels
#plt.rc('legend', fontsize=BIGGER_SIZE)    # legend fontsize
#plt.rc('figure', titlesize=BIGGER_SIZE)  # fontsize of the figure title


#====================PLOTTING INDUSTRY WEIGHTS====================

fig,ax = plt.subplots(figsize=(20,13))
plt.plot(industry_weights_ts)
plt.legend(industry_weights_ts.columns.tolist(),ncol=int(len(industry_names)/2),fontsize=15)

valsy = ax.get_yticks()
ax.set_yticklabels(['{:,.2%}'.format(y) for y in valsy])
ax.xaxis.set_major_formatter(mdates.DateFormatter("%b-%Y"))

plt.tight_layout()
if savefig:
    plt.savefig(r"Y:\Mobiliare\08 Finint Economia Reale Italia\00_Documenti_Reportistica\03 Comitati\Immagini update portafoglio\TS_weights.pdf",dpi=300)
plt.close(fig)

#=======================CORRELATION HEATMAP========================

ret_tot = pd.concat([eqt_ret,berk_ret],axis=1).iloc[1:,:].dropna(how='any',axis=1)
#corrHeat = fp.CorrHeatmap(pd.concat([eqt_ret,berk_ret],axis=1).iloc[1:,:].dropna(how='any',axis=1),savefig=savefig,ext = ext,loc = saveLoc)

eqt_corr = np.corrcoef(ret_tot,rowvar=False)
mask = np.zeros_like(eqt_corr)
mask[np.triu_indices_from(mask)] = True

f = plt.figure(figsize=(16,13))

with sns.axes_style("white"):
    ax = sns.heatmap(eqt_corr,mask=mask,vmax=1,vmin=-1,square =True,cmap="RdBu_r", \
                     xticklabels=ret_tot.columns.values,yticklabels=ret_tot.columns.values,center=0)
    
plt.title("Correlation Heatmap",fontdict = font)
if savefig:
    plt.savefig(r"Y:\Mobiliare\08 Finint Economia Reale Italia\00_Documenti_Reportistica\03 Comitati\Immagini update portafoglio\correlation.pdf",dpi=300)
plt.close(f)

#=======================RISK CONTRIBUTION========================
SD_cov = 20171229

eqt_last_w = eqt_w.iloc[-1,:][eqt_w.iloc[-1,:]>0]
eqt_ptf_ticker = eqt_last_w.index.tolist()
#eqt_VCV =eqt_ret[eqt_ret.columns.intersection(eqt_ptf_ticker)].cov().\
#        dropna(how='all',axis=1).dropna(how='all',axis=0)
eqt_VCV =eqt_ret[eqt_ret.columns.intersection(eqt_ptf_ticker)].cov(min_periods=30).\
        fillna(0)

eqt_sigma = (eqt_last_w.T @ eqt_VCV @ eqt_last_w )**0.5
eqt_MRC = (eqt_VCV @ eqt_last_w) / eqt_sigma
eqt_TRC = (eqt_MRC / eqt_sigma) * eqt_last_w

plt.figure(figsize=(16,13))
    
plt.bar(eqt_TRC.index,eqt_TRC, width = 0.4,color = 'darkblue')
plt.xticks(rotation = 90)
plt.title("Total Risk Contribution",fontdict = font)
if savefig:
    plt.savefig(r"Y:\Mobiliare\08 Finint Economia Reale Italia\00_Documenti_Reportistica\03 Comitati\Immagini update portafoglio\barchart_risk.pdf",dpi=300)

#=======================MARKET CAP BUBBLE CHART========================
eqt_last_w = eqt_w.iloc[-1,:][ eqt_w.iloc[-1,:]>0]
eqt_last_w = eqt_last_w.rename('w')

eqt_mark_cap = ptf_data[['TICKER','MKT_STOCKS']].dropna(how='all')
eqt_mark_cap.set_index('TICKER',inplace=True)

eqt_w_cap = pd.concat([eqt_last_w,eqt_mark_cap],axis=1,sort=False)
eqt_w_cap.sort_values('MKT_STOCKS',inplace=True)
eqt_w_cap = eqt_w_cap.dropna(how='any')
eqt_w_cap['MKT_STOCKS'] = eqt_w_cap['MKT_STOCKS'].div(1000000)

fig, ax = plt.subplots(figsize=(20,13))
p = ax.scatter(eqt_w_cap.index,eqt_w_cap['MKT_STOCKS'],\
            s=eqt_w_cap['w']*100*800,c=eqt_w_cap['w'],cmap="coolwarm", alpha=0.8, \
             linewidth=2)

plt.xticks(rotation=90)
ticks=np.arange(0,len(eqt_w_cap),1) 
ax.set_xticks(ticks)

m_i = next(x[0] for x in enumerate(eqt_w_cap['MKT_STOCKS']) if x[1]>500)
b_i = next(x[0] for x in enumerate(eqt_w_cap['MKT_STOCKS']) if x[1]>1000)

x = np.arange(np.size(eqt_w_cap,0))
y = eqt_w_cap['MKT_STOCKS']

ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.spines['bottom'].set_visible(False)
ax.spines['left'].set_visible(False)
plt.tick_params(axis='x', which='major', labelsize=13.5)

plt.plot([0,m_i-1],[-150,-150],'k',lw=0.1)
plt.plot([0,0],[-150,-100],'k',lw=0.1)
plt.plot([m_i-1,m_i-1],[-150,-100],'k',lw=0.1)

ax.annotate('<500Mln', xy=( m_i/2, -200), xytext=( m_i/2,-200), xycoords='data', 
            fontsize=8*1.5, ha='center', va='top',
            bbox=dict(boxstyle='square', fc='white'))

plt.plot([m_i,b_i-1],[-150,-150],'k',lw=0.1)
plt.plot([m_i,m_i],[-150,-100],'k',lw=0.1)
plt.plot([b_i-1,b_i-1],[-150,-100],'k',lw=0.1)

ax.annotate('500-1000 Mln', xy=( (b_i+m_i)/2, -200), xytext=( (b_i+m_i)/2,-200), xycoords='data', 
            fontsize=8*1.5, ha='center', va='top',
            bbox=dict(boxstyle='square', fc='white'))

plt.plot([b_i,np.size(eqt_w_cap,0)-1],[-150,-150],'k',lw=0.1)
plt.plot([b_i,b_i],[-150,-100],'k',lw=0.1)
plt.plot([np.size(eqt_w_cap,0)-1,np.size(eqt_w_cap,0)-1],[-150,-100],'k',lw=0.1)

ax.annotate('>1000 Mln', xy=( (b_i+np.size(eqt_w_cap,0)-1)/2, -200), xytext=( (b_i+np.size(eqt_w_cap,0)-1)/2,-200), xycoords='data', 
            fontsize=8*1.5, ha='center', va='top',
            bbox=dict(boxstyle='square', fc='white'))

for i,j in zip(x,y):
    rad = ((eqt_w_cap['w'][i]*100*800)**0.5/2)
    text = "%.2f%%"%(eqt_w_cap['w'][i]*100)
    ax.annotate(text, xy=(i,j+rad), xytext=(-17,rad), textcoords='offset points',fontsize=13.5)
    plt.plot([i,i],[0,eqt_w_cap['MKT_STOCKS'][i]-rad*4.8],'grey',linestyle='--', lw=0.1,dashes=(5, 5))
plt.tight_layout() 
if savefig:
    plt.savefig(r"Y:\Mobiliare\08 Finint Economia Reale Italia\00_Documenti_Reportistica\03 Comitati\Immagini update portafoglio\bubblechart.pdf",dpi=300)      

#=======================PIE CHART RISK ========================
colors = ['#ff9999','#66b3ff','#99ff99','#ffcc99','#657a97','#27a768','#faf771','#ff99cc']
m_df_eqt_ind = eqt_ind
m_df_eqt_ind.set_index('TICKER',inplace=True)

eqt_TRC_ind = pd.concat([pd.DataFrame(eqt_TRC,index=eqt_TRC.index),m_df_eqt_ind],axis=1,sort=False).dropna(how='any')
eqt_TRC_ind = pd.pivot_table(eqt_TRC_ind,0,'Industry Breakdown',aggfunc = np.sum)

exp = np.zeros_like(eqt_TRC_ind.index)
exp[np.array(eqt_TRC_ind.iloc[:,0]).argmax()] = 0.1

f,ax2 = plt.subplots(figsize=(14,6))
patches, texts, autotexts = ax2.pie(np.array(eqt_TRC_ind.iloc[:,0]),labels= eqt_TRC_ind.index,autopct='%1.1f ',explode=exp, \
                shadow = True,startangle=90,colors=colors)

ax2.set_title("By Risk Contribution\n",fontdict =  {'family': 'serif',
    'color':  'darkblue',
    'weight': 'normal',
    'size': 18,
    }) 
for i in range(len(texts)):
    texts[i].set_fontsize(15)
ax2.axis('equal')
if savefig:
    plt.savefig(r"Y:\Mobiliare\08 Finint Economia Reale Italia\00_Documenti_Reportistica\03 Comitati\Immagini update portafoglio\Industry_risk.pdf",dpi=300)
plt.close(f)

#=======================PIE CHART ALLOCATION ========================

eqt_CAP_ind = eqt_val.iloc[-1,:]
eqt_CAP_ind = eqt_CAP_ind[eqt_CAP_ind>0]
eqt_CAP_ind = eqt_CAP_ind.rename('cap')
eqt_CAP_ind = pd.concat([eqt_CAP_ind,m_df_eqt_ind],axis=1,sort=False).dropna(how='any')
eqt_CAP_ind = pd.pivot_table(eqt_CAP_ind,'cap','Industry Breakdown',aggfunc = np.sum)

exp = np.zeros_like(eqt_CAP_ind.index)
exp[np.array(eqt_CAP_ind.iloc[:,0]).argmax()] = 0.1

f,ax2 = plt.subplots(figsize=(14,6))
patches, texts, autotexts = ax2.pie(np.array(eqt_CAP_ind.iloc[:,0]),labels= eqt_CAP_ind.index,autopct='%1.1f ',explode=exp, \
                shadow = True,startangle=90,colors=colors)

ax2.set_title("By € Allocation\n",fontdict =  {'family': 'serif',
    'color':  'darkblue',
    'weight': 'normal',
    'size': 18,
    }) 
for i in range(len(texts)):
    texts[i].set_fontsize(15)
ax2.axis('equal')
if savefig:
    plt.savefig(r"Y:\Mobiliare\08 Finint Economia Reale Italia\00_Documenti_Reportistica\03 Comitati\Immagini update portafoglio\Industry_allocation.pdf",dpi=300)
plt.close(f)
print("="*60)    
print(" "*23+"END PROGRAM:")
print("="*60)   


if to_report:
    plt.style.use('fivethirtyeight')
    str_allocation = r'Y:\Mobiliare\08 Finint Economia Reale Italia\01_Front Office\02 Gestione\Scripts Python\allocation.CSV'
    pir_portfolio = pd.read_csv(str_allocation,encoding = "ISO-8859-1").dropna(how='all')
    pir_portfolio['ASSET CLASS'] = pir_portfolio['ASSET CLASS'].replace({'FIXED INCOME - CORPORATE':'FI Corp','FIXED INCOME - GOVT':'FI Govt'})
    col_list_pir_port = pir_portfolio.columns.tolist()
    pir_portfolio.rename(columns={'CPN_TYP':'CPN TYPE','MTY_YEARS_TDY':'MATURITY (Yrs)','DUR_ADJ_MTY_BID':'DURATION','Equity Market':'INDEX','Capitalization (Mln)':'CAP (Mln)'\
                                  ,'Div yield':'DVD YIELD','Target Price':'TARGET PRICE'},inplace=True)
                                  
    
    berk_price.to_excel(writer,'Benchmark price',startrow = 0,startcol=0,index=True)
    berk_ret.to_excel(writer,'Benchmark returns',startrow = 0,startcol=0,index=True)
    berk_perf.to_excel(writer,'Benchmark performance',startrow = 0,startcol=0,index=True)
    
    eqt_price.to_excel(writer,'Equity price',startrow = 0,startcol=0,index=True)
    eqt_ret.to_excel(writer,'Equity returns',startrow = 0,startcol=0,index=True)
    eqt_ptf_ret.to_excel(writer,'Equity performance',startrow = 0,startcol=0,index=True)
    eqt_perf.to_excel(writer,'Equity performance',startrow = 0,startcol=2,index=False)
    
    bond_val.to_excel(writer,'Bond Value',startrow = 0,startcol=0,index=True)
    bond_ptf_ret.to_excel(writer,'Bond performance',startrow = 0,startcol=0,index=True)
    bond_perf.to_excel(writer,'Bond performance',startrow = 0,startcol=2,index=False)
    
    pir_perf.to_excel(writer,'Pir performance',startrow = 0,startcol=0,index=True)
    pir_perf_a.to_excel(writer,'Pir performance',startrow = 0,startcol=3,index=False)
    
    w_asset_class.to_excel(writer,'Weights',startrow = 0,startcol=0,index=True)
    
    
#########################################################################
    #working on allocation
#########################################################################
    date_YTD = dt.date(2018,12,28)
    div_curryear_ret = pd.DataFrame(div_ret[div_ret.index>int(date_YTD.strftime("%Y%m%d"))].sum(axis=0),columns=['Div Ret'])

    ret_all = pd.concat([eqt_ptf_ret,bond_ptf_ret,pir_perf.loc[pir_perf.iloc[:,1].index.intersection(eqt_ptf_ret.index),['Return']]],axis=1)
    perf_ytd = (pir_portfolio.iloc[:,18].divide(pir_portfolio.iloc[:,15])-1)
    perf_mtd = (pir_portfolio.iloc[:,18].divide(pir_portfolio.iloc[:,16])-1)
    perf_weekly = (pir_portfolio.iloc[:,18].divide(pir_portfolio.iloc[:,17])-1)#.fillna("n.d.")
    last_price = pir_portfolio.iloc[:,18]
    for i in pir_portfolio.columns:
        try:
            int(i)
            pir_portfolio.drop([i], axis=1,inplace=True)
        except:
            pass
    
    pir_portfolio.insert(loc=10,column='Perf YTD',value = perf_ytd)
    pir_portfolio.insert(loc=11,column='Perf MTD',value = perf_mtd)
    pir_portfolio.insert(loc=12,column='Perf Weekly',value = perf_weekly)
    pir_portfolio.insert(loc=13,column='Last Price',value = last_price)
    pir_portfolio['CPN'] = pir_portfolio['CPN'].divide(100)
    pir_portfolio['CAP (Mln)'] = pir_portfolio['CAP (Mln)'].divide(1e06)
#    print(pir_portfolio.columns())
    bond_portfolio = pir_portfolio[pir_portfolio['ASSET CLASS'].str.contains('FI')]
    equity_portfolio = pir_portfolio[pir_portfolio['ASSET CLASS'].str.contains('EQUITY')]
    
    bond_portfolio.dropna(how='all',axis=1,inplace=True)
    bond_portfolio = bond_portfolio.sort_values('MKT VALUE',ascending=False)
    equity_portfolio.dropna(how='all',axis=1,inplace=True)
    equity_portfolio=equity_portfolio.sort_values('MKT VALUE',ascending=False)
    
    equity_portfolio = pd.merge( equity_portfolio,div_curryear_ret, right_index=True, left_on='TICKER')
    equity_portfolio.to_excel(writer,'Equity port',startrow = 4,startcol=2,index=False,na_rep="n.d.")
    bond_portfolio.to_excel(writer,'Bond port',startrow = 4,startcol=2,index=False,na_rep="n.d.") 
    
    best_pf = []
    worst_pf = []
    
    for i in [equity_portfolio,bond_portfolio]:
        for j in ['Perf Weekly','Perf MTD','Perf YTD']:
            idx_m = i[j].idxmax()
            idx_w = i[j].idxmin()
            val_max = float(i[j].max())
            val_min = float(i[j].min())
            if any(i['ASSET CLASS']=='EQUITY'):
                to_app_max = [i.loc[idx_m,'TICKER'],val_max,'EQUITY B']
                to_app_min = [i.loc[idx_w,'TICKER'],val_min,'EQUITY W']
            else:
                to_app_max = [i.loc[idx_m,'SECURITY NAME'],val_max,'BOND B']
                to_app_min = [i.loc[idx_w,'SECURITY NAME'],val_min,'BOND W']
            best_pf.append(to_app_max)
            worst_pf.append(to_app_min)
      
    idx_df = ['Perf Weekly','Perf MTD','Perf YTD']*2
    col_df = ['Name','Perf','AC']
    best_pf = pd.DataFrame(best_pf,index=idx_df,columns=col_df)
    worst_pf = pd.DataFrame(worst_pf,index=idx_df,columns=col_df)
    best_worst = pd.concat([best_pf,worst_pf])

    #average PB and DE -  EQUITY
    ints_pb = equity_portfolio.index.intersection(equity_portfolio['P/B'].dropna().index)
    avg_pb = (equity_portfolio.loc[ints_pb,'MKT VALUE']/(equity_portfolio.loc[ints_pb,'MKT VALUE'].sum()).T @ equity_portfolio['P/B'].dropna())  
    ints_de = equity_portfolio.index.intersection(equity_portfolio['D/E'].dropna().index)
    avg_de = (equity_portfolio.loc[ints_de,'MKT VALUE']/(equity_portfolio.loc[ints_de,'MKT VALUE'].sum()).T @ equity_portfolio['D/E'].dropna())  
    eqt_max_all =[equity_portfolio.loc[equity_portfolio['MKT VALUE'].idxmax(),'TICKER'],equity_portfolio['MKT VALUE'].max()]
    eqt_cap_avg = equity_portfolio['CAP (Mln)'].mean()
    eqt_cap_median = equity_portfolio['CAP (Mln)'].median()
    #average maturity and duration - BOND
    bond_max_all =[bond_portfolio.loc[bond_portfolio['MKT VALUE'].idxmax(),'SECURITY NAME'],bond_portfolio['MKT VALUE'].max()]
    ints_dur = bond_portfolio.index.intersection(bond_portfolio['DURATION'].dropna().index)
    avg_dur = (bond_portfolio.loc[ints_dur,'MKT VALUE']/(bond_portfolio.loc[ints_dur,'MKT VALUE'].sum())).T @ bond_portfolio['DURATION'].dropna()  
    ints_mat = bond_portfolio.index.intersection(bond_portfolio['MATURITY (Yrs)'].dropna().index)
    avg_mat = (bond_portfolio.loc[ints_mat,'MKT VALUE']/(bond_portfolio.loc[ints_mat,'MKT VALUE'].sum()).T @ bond_portfolio['MATURITY (Yrs)'].dropna())
    bond_max_all =[bond_portfolio.loc[bond_portfolio['MKT VALUE'].idxmax(),'ISSUER'],bond_portfolio['MKT VALUE'].max()]
    
    weekly_perf_all = df_perf_all.iloc[-1,:]/df_perf_all.iloc[-6,:]-1
    
    weekly_std_all = ret_all.iloc[-5:,:].std()*256**0.5
    ytd_std_all =  ret_all.loc[int(date_YTD.strftime('%Y%m%d')):,:].std()*256**0.5
    mtd_std_all =  ret_all.loc[int(last_m):,:].std()*256**0.5
    n_pos_months = df_monthly_perf[df_monthly_perf>0].count()/np.size(df_monthly_perf,0)
    #equity 
    equity_list_1 = [np.size(equity_portfolio,0),equity_portfolio['MKT VALUE'].sum(),'',\
                     weekly_perf_all['port_eqt_ret'],perf_summary.loc['Equity','MTD']/100,perf_summary.loc['Equity','YTD']/100,'',\
                     weekly_std_all['port_eqt_ret'],mtd_std_all['port_eqt_ret'],ytd_std_all['port_eqt_ret'],\
                     '',n_pos_months['Equity Monthly Ret']]
    
    equity_list_2 = pd.DataFrame(pd.concat([pd.DataFrame(eqt_max_all,index=['Name','Perf']).T,best_worst.iloc[:,:-1]])).iloc[:4,:]
    equity_list_2p2 = best_worst.iloc[6:9,:-1]
    equity_list_3 = pd.DataFrame([[avg_pb],[avg_de],[eqt_cap_median],[eqt_cap_avg]])
    
    equity_list_1  = pd.DataFrame(equity_list_1)
    
    equity_list_1.to_excel(writer,'EQUITY Chart',startrow = 6,startcol=3,index=False,header =False,na_rep="n.d.")
    equity_list_2.iloc[0,:].T.to_excel(writer,'EQUITY Chart',startrow = 6,startcol=7,index=False,header =False,na_rep="n.d.")
    equity_list_2.iloc[1:,:].to_excel(writer,'EQUITY Chart',startrow = 9,startcol=7,index=False,header =False,na_rep="n.d.")

    
    equity_list_2p2.to_excel(writer,'EQUITY Chart',startrow = 13,startcol=7,index=False,header =False,na_rep="n.d.")
    equity_list_3.to_excel(writer,'EQUITY Chart',startrow = 19,startcol=7,index=False,header =False,na_rep="n.d.")
    #bond 
   
    bond_list_1 = [np.size(bond_portfolio,0),bond_portfolio['MKT VALUE'].sum(),'',\
                 weekly_perf_all['port_bond_ret'],perf_summary.loc['Bond','MTD']/100,perf_summary.loc['Bond','YTD']/100,'',\
                 weekly_std_all['port_bond_ret'],mtd_std_all['port_bond_ret'],ytd_std_all['port_bond_ret'],\
                 '',n_pos_months['Bond Monthly Ret']]
    bond_list_2 = pd.DataFrame(pd.concat([pd.DataFrame(bond_max_all,index=['Name','Perf']).T,best_worst.iloc[:,:-1].iloc[3:6,:]]))
    bond_list_2p2 = best_worst.iloc[9:,:-1]
    bond_list_3 = pd.DataFrame([[avg_dur],[avg_mat]])

    pd.DataFrame(bond_list_1).to_excel(writer,'BOND Chart',startrow = 6,startcol=3,index=False,header =False,na_rep="n.d.")
    
    bond_list_2.iloc[0,:].T.to_excel(writer,'BOND Chart',startrow = 6,startcol=7,index=False,header =False,na_rep="n.d.")
    bond_list_2.iloc[1:,:].to_excel(writer,'BOND Chart',startrow = 9,startcol=7,index=False,header =False,na_rep="n.d.")
    bond_list_2p2.to_excel(writer,'BOND Chart',startrow = 13,startcol=7,index=False,header =False,na_rep="n.d.")
    bond_list_3.to_excel(writer,'BOND Chart',startrow = 19,startcol=7,index=False,header =False,na_rep="n.d.")
    #fund
    fund_list_1 = [np.size(pir_portfolio,0),pir_perf.iloc[-1,0],'',\
                 weekly_perf_all['Return'],perf_summary.loc['PIR','MTD']/100,perf_summary.loc['PIR','YTD']/100,'',\
                 weekly_std_all['Return'],mtd_std_all['Return'],ytd_std_all['Return'],\
                 '',n_pos_months['PIR Monthly Ret']]
    fund_list_2 =[[pir_perf.iloc[-1,2]],[pir_portfolio.loc[np.size(pir_portfolio,0)-1,'MKT VALUE']]]
    fund_list_3 = [[pir_perf.iloc[-5,0]],[pir_perf.loc[int(last_m),'Share']],[pir_perf.loc[int(date_YTD.strftime('%Y%m%d')),'Share']]]
    fund_list_4 = [[pir_perf.iloc[-5,2]],[pir_perf.loc[int(last_m),'Nav']],[pir_perf.loc[int(date_YTD.strftime('%Y%m%d')),'Nav']]]

    pd.DataFrame(fund_list_1).to_excel(writer,'PIR Chart',startrow = 6,startcol=3,index=False,header =False,na_rep="n.d.")
    pd.DataFrame(fund_list_2).to_excel(writer,'PIR Chart',startrow = 6,startcol=7,index=False,header =False,na_rep="n.d.")
    pd.DataFrame(fund_list_3).to_excel(writer,'PIR Chart',startrow = 9,startcol=7,index=False,header =False,na_rep="n.d.")
    pd.DataFrame(fund_list_4).to_excel(writer,'PIR Chart',startrow = 13,startcol=7,index=False,header =False,na_rep="n.d.")
     

#########################################################################
   #re read
#########################################################################     
    
    pir_portfolio = pd.read_csv(str_allocation,encoding = "ISO-8859-1").dropna(how='all')
    pir_portfolio['ASSET CLASS'] = pir_portfolio['ASSET CLASS'].replace({'FIXED INCOME - CORPORATE':'FI Corp','FIXED INCOME - GOVT':'FI Govt'})
    
    bond_portfolio = pir_portfolio[pir_portfolio['ASSET CLASS'].str.contains('FI')]
    equity_portfolio = pir_portfolio[pir_portfolio['ASSET CLASS'].str.contains('EQUITY')]
    
    
#########################################################################
   #fund statistics
#########################################################################  
    
    
    sr_all = (ret_all.mean()-rf)/ret_all.std()
    sr_all_mean = ret_all.mean()*256
    sr_all_std = ret_all.std()*256**0.5
    
    summary_data_pir = [int(pir_perf.index[-1]),pir_perf.iloc[-1,0],pir_perf.iloc[-1,2],pir_portfolio.loc[pir_portfolio.index[-1],'MKT VALUE'],\
                                perf_summary.loc['PIR','YTD']/100,perf_summary.loc['PIR','MTD']/100,pir_perf.iloc[-1,0]/500-1]
    summary_pir = pd.DataFrame(summary_data_pir,index=['Date','NAV/Share','NAV','Cash','YTD','MTD','Inception',],columns=['PIR'])
    
    fund_fact_pir = [['IT0005261125','FIERITA IM',pir_perf.iloc[-1,0],pir_perf.iloc[-1,2]],['IT0005273575','FIERPIR IM',pir_perf.iloc[-1,0],pir_perf.iloc[-1,2]]]
    fund_fact_pir = pd.DataFrame(np.stack(fund_fact_pir),columns=['ISIN','Bloomberg','Share Price','NAV'],index=['Class A Share','Class PIR Share'])
    
    X_reg = np.vstack(([np.ones_like(berk_ret['FTSEMIB Index'])],berk_ret['FTSEMIB Index'].tolist())).T
    beta_fund = (np.linalg.inv(X_reg.T @ X_reg)) @ (X_reg.T @ pir_perf.loc[pir_perf.iloc[:,1].index.intersection(eqt_ptf_ret.index),['Return']])
    
    beta_equity = (np.linalg.inv(X_reg.T @ X_reg)) @ (X_reg.T @ eqt_ptf_ret)
    
    
    beta_bond = (np.linalg.inv(X_reg.T @ X_reg)) @ (X_reg.T @ bond_ptf_ret)
    try:
        beta_fund = beta_fund.iloc[1,0]
        beta_equity = beta_equity.iloc[1,0]
        beta_bond = beta_bond.iloc[1,0]
    except: 
        beta_fund = beta_fund[1,0]
        beta_equity = beta_equity[1,0]
        beta_bond = beta_bond[1,0]
    
    stats_fund =[sr_all_mean[2],sr_all_std[2],sr_all[2],beta_fund,md[2],df_monthly_perf.iloc[:,2].idxmax(),df_monthly_perf.iloc[:,2].idxmin()]
    stats_fund = pd.DataFrame(stats_fund,index=['Annualized return','Annualized std','SR','Beta (FTSEMIB)','Max Drawdown','Best Month','Worst Month'],columns=['Fund'])

    stats_equity =[sr_all_mean[0],sr_all_std[0],sr_all[0],beta_equity,md[0],df_monthly_perf.iloc[:,0].idxmax(),df_monthly_perf.iloc[:,0].idxmin()]
    stats_equity = pd.DataFrame(stats_equity,index=['Annualized return','Annualized std','SR','Beta (FTSEMIB)','Max Drawdown','Best Month','Worst Month'],columns=['Equity'])

    stats_bond =[sr_all_mean[1],sr_all_std[1],sr_all[1],beta_bond,md[1],df_monthly_perf.iloc[:,1].idxmax(),df_monthly_perf.iloc[:,1].idxmin()]
    stats_bond = pd.DataFrame(stats_bond,index=['Annualized return','Annualized std','SR','Beta (FTSEMIB)','Max Drawdown','Best Month','Worst Month'],columns=['Bond'])
    
    len_index = dt.date.today().year -2017+1
    year_row = [dt.date.today().year-i+1 for i in range(len_index,0,-1)]
    col = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec','YTD']
    df_equity_perf = pd.DataFrame(np.zeros((len_index,13)),index=year_row,columns=col)
    df_bond_perf = pd.DataFrame(np.zeros((len_index,13)),index=year_row,columns=col)
    df_fund_perf = pd.DataFrame(np.zeros((len_index,13)),index=year_row,columns=col)
    
    for j in range(3):
        for i in range(13):
            try:
                idx_name = df_equity_perf.columns[i]+' ' +str(df_equity_perf.index[j])
                df_equity_perf.iloc[j,i] = float(df_monthly_perf.loc[idx_name][0])
                df_bond_perf.iloc[j,i] = df_monthly_perf.loc[idx_name][1]
                df_fund_perf.iloc[j,i] = df_monthly_perf.loc[idx_name][2]
            except:
                df_equity_perf.iloc[j,i] = np.nan
                df_bond_perf.iloc[j,i] = np.nan
                df_fund_perf.iloc[j,i] = np.nan
            
    df_equity_perf['YTD'] = np.array([eqt_2017,eqt_2018,eqt_ytd]) 
    df_bond_perf['YTD'] = np.array([bond_2017,bond_2018,bond_ytd])
    df_fund_perf['YTD'] = np.array([pir_2017,pir_2018,pir_ytd])
    
    df_equity_perf.to_excel(writer,'Performance sheet',startrow = 3,startcol=2,index_label = 'Equity',index=True)
    df_bond_perf.to_excel(writer,'Performance sheet',startrow = 3,startcol=17,index_label = 'Bond',index=True)
    df_fund_perf.to_excel(writer,'Performance sheet',startrow = 3,startcol=32,index_label = 'Fund',index=True)
    
    date_timestamp  = pd.to_datetime(eqt_ptf_ret.index, format='%Y%m%d')
    date_last = date_timestamp[-1]

    date_1m = date_last - BDay(20) #20 working days = 1 month
    date_3m = date_last - BDay(60) #60 working days = 3 month
    date_6m = date_last - BDay(120) #60 working days = 3 month
    date_1yr = date_last - BDay(260) #60 working days = 3 month
    
    date_1m = int(date_1m.strftime('%Y%m%d'))
    date_3m = int(date_3m.strftime('%Y%m%d'))#date_6m = 
    date_6m = int(date_6m.strftime('%Y%m%d'))
    date_1yr = int(date_1yr.strftime('%Y%m%d'))
    
    ret_all_1m = ret_all.loc[date_1m:,:]
    ret_all_3m = ret_all.loc[date_3m:,:]
    ret_all_6m = ret_all.loc[date_6m:,:]
    ret_all_1yr = ret_all.loc[date_1yr:,:]
    
    cum_ret1m = ((ret_all_1m.add(1).cumprod()).iloc[-1,:]-1).tolist()
    cum_ret3m = ((ret_all_3m.add(1).cumprod()).iloc[-1,:]-1).tolist()
    cum_ret6m = ((ret_all_6m.add(1).cumprod()).iloc[-1,:]-1).tolist()
    cum_ret1yr =((ret_all_1yr.add(1).cumprod()).iloc[-1,:]-1).tolist()
    
    sr_1m = ((ret_all_1m.mean()-rf)/ret_all_1m.std()).tolist()
    sr_3m = ((ret_all_3m.mean()-rf)/ret_all_3m.std()).tolist()
    sr_6m = ((ret_all_6m.mean()-rf)/ret_all_6m.std()).tolist()
    sr_1yr = ((ret_all_1yr.mean()-rf)/ret_all_1yr.std()).tolist()
    
    avg_1m = (ret_all_1m.mean()*256).tolist()
    avg_3m = (ret_all_3m.mean()*256).tolist()
    avg_6m = (ret_all_6m.mean()*256).tolist()
    avg_1yr = (ret_all_1yr.mean()*256).tolist()
    
    std_1m = (ret_all_1m.std()*256**0.5).tolist()
    std_3m = (ret_all_3m.std()*256**0.5).tolist()
    std_6m = (ret_all_6m.std()*256**0.5).tolist()
    std_1yr = (ret_all_1yr.std()*256**0.5).tolist()
    
    roll_analysis = np.stack((cum_ret1m,cum_ret3m,cum_ret6m,cum_ret1yr,sr_1m,sr_3m,sr_6m,sr_1yr,avg_1m,avg_3m,avg_6m,avg_1yr,\
                             std_1m,std_3m,std_6m,std_1yr))
    
    index_roll = ['Cum Return(1m)','Cum Return(3m)','Cum Return(6m)','Cum Return(1yr)',\
                  'SR (1m)','SR (3m)','SR (6m)','SR (1yr)',\
                  'mean Return (1m)','mean Return (3m)','mean Return (6m)','mean Return (1yr)',\
                  'volatility (1m)','volatility (3m)','volatility (6m)','volatility (1yr)',]
    roll_analysis = pd.DataFrame(roll_analysis,columns=['Equity','Bond','Fund'],)
    
    roll_analysis.loc[:3,['Equity']].to_excel(writer,'Performance analysis',startrow = 24,startcol=5,index = False,header=False)
    roll_analysis.loc[4:7,['Equity']].to_excel(writer,'Performance analysis',startrow = 30,startcol=5,index = False,header=False)
    roll_analysis.loc[8:11,['Equity']].to_excel(writer,'Performance analysis',startrow = 36,startcol=5,index = False,header=False)
    roll_analysis.loc[12:15,['Equity']].to_excel(writer,'Performance analysis',startrow = 42,startcol=5,index = False,header=False)

    roll_analysis.loc[:3,['Bond']].to_excel(writer,'Performance analysis',startrow = 24,startcol=8,index = False,header=False)
    roll_analysis.loc[4:7,['Bond']].to_excel(writer,'Performance analysis',startrow = 30,startcol=8,index = False,header=False)
    roll_analysis.loc[8:11,['Bond']].to_excel(writer,'Performance analysis',startrow = 36,startcol=8,index = False,header=False)
    roll_analysis.loc[12:15,['Bond']].to_excel(writer,'Performance analysis',startrow = 42,startcol=8,index = False,header=False)

    roll_analysis.loc[:3,['Fund']].to_excel(writer,'Performance analysis',startrow = 24,startcol=2,index = False,header=False)
    roll_analysis.loc[4:7,['Fund']].to_excel(writer,'Performance analysis',startrow = 30,startcol=2,index = False,header=False)
    roll_analysis.loc[8:11,['Fund']].to_excel(writer,'Performance analysis',startrow = 36,startcol=2,index = False,header=False)
    roll_analysis.loc[12:15,['Fund']].to_excel(writer,'Performance analysis',startrow = 42,startcol=2,index = False,header=False)
    
    roll_analysis.set_index([index_roll],inplace =True)
    roll_analysis.to_excel(writer,'Performance sheet',startrow = 15,startcol=2,index_label = 'Name',index=True) 
#########################################################################
    #SUBSCRIPTIONS
#########################################################################   
    str_file_wdraws = 'refunds.txt'
    str_file_subs = 'subs.txt'
    
    subs = pd.read_csv(str_file_subs,sep='\t',header=0,index_col=None)
    wdraws = pd.read_csv(str_file_wdraws,sep='\t',header=0,index_col=None)
    
    cols_subs = [x for x in subs.columns.tolist() if not ('Unnamed' in x)]
    subs = subs.dropna(how='all',axis=1)
    subs.columns = cols_subs
    subs.iloc[:,0] = subs.iloc[:,0].apply(np.int)
    subs.set_index(subs.iloc[:,0].name,inplace=True)
    
    cols_wdraws = [x for x in wdraws.columns.tolist() if not ('Unnamed' in x)]
    wdraws = wdraws.dropna(how='all',axis=1)
    wdraws.columns = cols_wdraws
    wdraws.iloc[:,0] = wdraws.iloc[:,0].apply(np.int)
    wdraws.set_index(wdraws.iloc[:,0].name,inplace=True)
    
    subs = subs.subtract(5e6).diff()
    
    wdraws = wdraws.reindex(eqt_ptf_ret.index).fillna(0)
    subs = subs.reindex(eqt_ptf_ret.index).fillna(0)
    
    net_subs = subs['SUBS'] #- wdraws['WITHDRAWALS']
    net_subs = net_subs.to_frame('NET SUBS')
    
    
    net_subs.set_index(date_timestamp.strftime('%b-%y'),inplace=True)
    net_subs = pd.pivot_table(net_subs,'NET SUBS',net_subs.index,aggfunc=np.sum)
    net_subs['sort_date'] = pd.to_datetime(net_subs.index,format = '%b-%y')
    net_subs.sort_values(by='sort_date',inplace=True)
    net_subs.drop('sort_date',axis=1,inplace=True)
    #SUM BACK wdrawals seed money (april and may 2018)
    net_subs.loc['Apr-18'] = net_subs.loc['Apr-18'] + 1.00968905e6
    net_subs.loc['May-18'] = net_subs.loc['May-18'] + .509128e6
    cum_subs = net_subs.cumsum()
    
    
#########################################################################
    #SUBS PASTE IN SHEET
#########################################################################  
    
    net_subs.to_excel(writer,'Subs',startrow = 4,startcol=2,index = True,header=True,index_label='Month')
    cum_subs = net_subs.cumsum().to_excel(writer,'Subs',startrow = 4,startcol=5,index = True,header=True,index_label='Month')
    
#########################################################################
    #READ TOTAL ASSETS FILE
#########################################################################      
    
    data_txt = pd.read_csv('file.txt',delimiter='\t')
    
    
    columns = data_txt.columns.tolist()
    
    columns = [x for x in columns if 'Unnamed' not in x]
    data_txt.dropna(how='all',axis=1,inplace=True)
    data_txt.columns = columns
    
    date_read = [str(int(x)) for x in data_txt[columns[0]]]
    date_fl = pd.to_datetime([str(int(x)) for x in data_txt[columns[0]]],format='%Y/%m/%d',yearfirst =True)
    date = [dt.datetime.strftime(x,"%b-%y") for x in date_fl]
                         
    data_txt.set_index([date],inplace=True)
    data_txt.drop(columns[0],inplace=True,axis=1)
    
    tot_assets = data_txt.sum(1)
    perc = data_txt.divide(tot_assets,axis=0)
#########################################################################
    #pivot tables
#########################################################################   
    
    fund_industry = double_pivot(pir_portfolio,'MKT VALUE','INDUSTRY SECTOR').sort_values('MKT VALUE',ascending=False)
    fund_subindustry = double_pivot(pir_portfolio,'MKT VALUE','INDUSTRY GROUP').sort_values('MKT VALUE',ascending=False)
    fund_ac = double_pivot(pir_portfolio,'MKT VALUE','ASSET CLASS').sort_values('MKT VALUE',ascending=False)
    fund_listed = double_pivot(pir_portfolio,'MKT VALUE','LISTED').sort_values('MKT VALUE',ascending=False)
    fund_listed.rename({0:"unlisted",1:"listed"},inplace=True)
    fund_country = double_pivot(pir_portfolio,'MKT VALUE','Country').sort_values('MKT VALUE',ascending=False)

    
    func = np.vectorize(cond)
    func1 = np.vectorize(cond1)
    func_div = np.vectorize(cond_div)
    func_cap = np.vectorize(cond_cap)
    
    maturity = func(bond_portfolio['MTY_YEARS_TDY'])
    duration = func1(bond_portfolio['DUR_ADJ_MTY_BID'])
    div = func_div(equity_portfolio['Div yield'])
    cap = func_cap(equity_portfolio['Capitalization (Mln)'])
    
    bond_portfolio['MATURITY'] = maturity
    bond_portfolio['DURATION'] = duration
    equity_portfolio['DIVIDEND'] = div
    equity_portfolio['CAP'] = cap
    
    bond_industry = double_pivot(bond_portfolio,'MKT VALUE','INDUSTRY SECTOR').sort_values('MKT VALUE',ascending=False)
    bond_ranking =double_pivot(bond_portfolio,'MKT VALUE','RANKING').sort_values('MKT VALUE',ascending=False)
    bond_maturity = double_pivot(bond_portfolio,'MKT VALUE','MATURITY').sort_values('MKT VALUE',ascending=False)
    bond_duration = double_pivot(bond_portfolio,'MKT VALUE','DURATION').sort_values('MKT VALUE',ascending=False)
    bond_coupon = double_pivot(bond_portfolio,'MKT VALUE','CPN_TYP').sort_values('MKT VALUE',ascending=False)
    bond_country = double_pivot(bond_portfolio,'MKT VALUE','Country').sort_values('MKT VALUE',ascending=False)
    
    equity_industry = double_pivot(equity_portfolio,'MKT VALUE','INDUSTRY SECTOR').sort_values('MKT VALUE',ascending=False)
    equity_ranking = double_pivot(equity_portfolio,'MKT VALUE','RANKING').sort_values('MKT VALUE',ascending=False)
    equity_market = double_pivot(equity_portfolio,'MKT VALUE','Equity Market').sort_values('MKT VALUE',ascending=False)
    equity_dividend = double_pivot(equity_portfolio,'MKT VALUE','DIVIDEND').sort_values('MKT VALUE',ascending=False)
    equity_cap = double_pivot(equity_portfolio,'MKT VALUE','CAP').sort_values('MKT VALUE',ascending=False)
    equity_country = double_pivot(equity_portfolio,'MKT VALUE','Country').sort_values('MKT VALUE',ascending=False)
        
#########################################################################
    #pivot tables END
#########################################################################    
    #paste asset class security type and ranking
    skip_row = 0
    
    stats_fund.to_excel(writer,'PIR Chart',startrow = 19,startcol=3,index = False,header=False)
    stats_equity.to_excel(writer,'EQUITY Chart',startrow = 19,startcol=3,index = False,header=False)
    stats_bond.to_excel(writer,'BOND Chart',startrow = 19,startcol=3,index = False,header=False)
    
    stats_fund.to_excel(writer,'Summary',startrow = 3+skip_row,startcol=2,index = True,header=True,index_label='Stats')
    stats_equity.to_excel(writer,'Summary',startrow = 3+skip_row,startcol=6,index = True,header=True,index_label='Stats')
    stats_bond.to_excel(writer,'Summary',startrow = 3+skip_row,startcol=10,index = True,header=True,index_label='Stats')
    
    skip_row += np.size(stats_fund,0)+2
    
    fund_ac.to_excel(writer,'Summary',startrow = 3+skip_row,startcol=2,index=True)
    equity_ranking.to_excel(writer,'Summary',startrow = 3+skip_row,startcol=6,index=True)
    bond_ranking.to_excel(writer,'Summary',startrow = 3+skip_row,startcol=10,index=True)
    
    skip_row += max(np.size(equity_ranking,0),np.size(bond_ranking,0),np.size(fund_ac,0))+2
    #paste country breakdown
    fund_country.to_excel(writer,'Summary',startrow = 3+skip_row,startcol=2,index=True)
    equity_country.to_excel(writer,'Summary',startrow = 3+skip_row,startcol=6,index=True)
    bond_country.to_excel(writer,'Summary',startrow = 3+skip_row,startcol=10,index=True)
  
    skip_row += max(np.size(equity_country,0),np.size(bond_country,0),np.size(fund_country,0))+2
    
    #paste industry breakdown
    
    fund_industry.to_excel(writer,'Summary',startrow = 3+skip_row,startcol=2,index=True)
    equity_industry.to_excel(writer,'Summary',startrow = 3+skip_row,startcol=6,index=True)
    bond_industry.to_excel(writer,'Summary',startrow = 3+skip_row,startcol=10,index=True)

    skip_row += max(np.size(equity_industry,0),np.size(bond_industry,0),np.size(fund_industry,0))+2

    #paste listed dividend duration
    
    fund_listed.to_excel(writer,'Summary',startrow = 3+skip_row,startcol=2,index=True)
    equity_dividend.to_excel(writer,'Summary',startrow = 3+skip_row,startcol=6,index=True)
    bond_duration.to_excel(writer,'Summary',startrow = 3+skip_row,startcol=10,index=True)
 #########################################################################
    #paste in sheet
#########################################################################   
    skip_row = 24
    
    equity_ranking.to_excel(writer,'EQUITY Chart',startrow = 3+skip_row,startcol=2,index=True)
    equity_industry.to_excel(writer,'EQUITY Chart',startrow = 3+skip_row,startcol=6,index=True)
    
    skip_row += max(np.size(equity_ranking,0),np.size(equity_industry,0))+2
    equity_dividend.to_excel(writer,'EQUITY Chart',startrow = 3+skip_row,startcol=2,index=True)
    equity_country.to_excel(writer,'EQUITY Chart',startrow = 3+skip_row,startcol=6,index=True,index_label = "ISSUER'S COUNTRY")
    
    skip_row = 24
    bond_ranking.to_excel(writer,'BOND Chart',startrow = 3+skip_row,startcol=2,index=True)
    bond_industry.to_excel(writer,'BOND Chart',startrow = 3+skip_row,startcol=6,index=True)
    skip_row += max(np.size(bond_ranking,0),np.size(bond_industry,0))+2

    bond_duration.to_excel(writer,'BOND Chart',startrow = 3+skip_row,startcol=2,index=True)
    bond_country.to_excel(writer,'BOND Chart',startrow = 3+skip_row,startcol=6,index=True,index_label = "ISSUER'S COUNTRY")

    skip_row = 24
    fund_ac.to_excel(writer,'PIR Chart',startrow = 3+skip_row,startcol=2,index=True)
    fund_industry.to_excel(writer,'PIR Chart',startrow = 3+skip_row,startcol=6,index=True)
    skip_row += max(np.size(fund_ac,0),np.size(fund_industry,0))+2
    
    fund_listed.to_excel(writer,'PIR Chart',startrow = 3+skip_row,startcol=2,index=True)
    fund_country.to_excel(writer,'PIR Chart',startrow = 3+skip_row,startcol=6,index=True,index_label = "ISSUER'S COUNTRY")
#=====================================================
#VAR
#=====================================================
    pir = pir_perf[['Return']]*100
    date_index = pir.index.tolist()
    date_index = [dt.datetime.strptime(str(int(x)),'%Y%m%d').strftime('%Y-%m-%d') for x in date_index]
    pir.set_index([np.array(date_index,dtype=np.datetime64)],inplace=True)
    
    split_date = '2018-04-30'
    dist = 'normal'
    am = arch.arch_model(pir,vol='GARCH',p=1,o=0,q=1,dist=dist)
    #recursive forecast
    index = pir.index
    sl = 0
    el = np.where(index >= split_date)[0].min()
    
    forecasts_s2 = {}
    forecasts_mean = {}
    q_var = {}
    q_filtered ={}
    for i in range(np.size(pir,0)-el+1):
        
        pir_slice = pir[:i+el]
        
        ##for parametric var
        res = am.fit(disp='off',last_obs=i+el)
        temp_s2 = res.forecast(horizon=1).variance
        temp_mean = res.forecast(horizon=1).mean
        
        fcast_s2 = temp_s2.iloc[i+el-1]
        fcast_mean = temp_mean.iloc[i+el-1]
        
        forecasts_s2[fcast_s2.name] = fcast_s2
        forecasts_mean[fcast_mean.name] = fcast_mean
        if dist == 'normal':
            q = np.array([-2.33,-1.64])
        else:
            q = am.distribution.ppf(np.array([0.01,0.05]), res.params[-2:])
        q_var[fcast_mean.name] = q
        
        ##for non parametric var
        std_rets = (pir_slice - res.params['mu']).div(res.conditional_volatility,axis=0)
        std_rets = std_rets.dropna()
        q_filt = std_rets.quantile([.01, .05])
        q_filt =[x[0] for x in q_filt.values]
        q_filtered[fcast_mean.name] = q_filt
    
    
    cond_mean = pd.DataFrame.from_dict(forecasts_mean).T  
    cond_var = pd.DataFrame.from_dict(forecasts_s2).T 
    quantiles =  pd.DataFrame.from_dict(q_var).T 
    quantiles_empirical =  pd.DataFrame.from_dict(q_filtered).T 
    
    
    value_at_risk = -cond_mean.values - np.sqrt(cond_var).values * quantiles
    value_at_risk.columns = ['1%', '5%']
    value_at_risk = value_at_risk[split_date:]
    
    value_at_risk_empirical = -cond_mean.values - np.sqrt(cond_var).values * quantiles_empirical
    value_at_risk_empirical.columns = ['1%', '5%']
    value_at_risk_empirical = value_at_risk_empirical[split_date:]
    
    next_res = am.fit(disp='off')
    
    next_forecast = next_res.forecast(horizon=5,method = 'simulation',simulations=1000)
    next_cond_mean = next_forecast.mean.dropna()
    next_cond_var = next_forecast.variance.dropna()
    
    if dist == 'normal':
        next_q = np.array([-2.33,-1.64])  
    else:
        next_q = am.distribution.ppf(np.array([0.01,0.05]), next_res.params[-2:])
    next_value_at_risk = -next_cond_mean.values - np.sqrt(next_cond_var).values * next_q[None, :].T
    next_value_at_risk = pd.DataFrame(
        next_value_at_risk, index=['1%', '5%'], columns=next_cond_var.columns)   
    
    next_std_rets = (pir- res.params['mu']).div(next_res.conditional_volatility,axis=0)
    next_std_rets = next_std_rets.dropna()
    next_quantiles = next_std_rets.quantile([.01, .05])
    
    am_mc = arch.arch_model(pir,vol='GARCH',p=1,o=0,q=1,dist='normal')
    mc_next_res = am_mc.fit(disp='off')
    garch_coef = mc_next_res.params.values
    simga2_t = pir.var()
    sigm2_t_1 = garch_coef[1] + garch_coef[2] * pir.iloc[-1,0]**2 + garch_coef[3] * simga2_t
    sigma_t_1 = sigm2_t_1**0.5
    
    n_obs = 10000
    
    sim_ret = np.zeros((n_obs,5))
    for i in range(n_obs):
        for j in range(5):
            if j == 0:
                sigm2_updated = sigma_t_1
            draw = np.random.normal()
            r_hat =  (sigma_t_1 * draw).values[0] + garch_coef[0]
            
            sim_ret[i,j] = r_hat
            sigm2_updated = garch_coef[1] + garch_coef[2] * r_hat**2 + garch_coef[3] * sigm2_updated
    
    sim_ret = pd.DataFrame(sim_ret)
    mc_value_at_risk = sim_ret.quantile([.01, .05],axis=0)
    mc_value_at_risk.index = ['1%', '5%']
    mc_value_at_risk.columns = ['h.'+str(i) for i in range(1,6)]    
    
    
#    next_quantiles = [x[0] for x in next_quantiles.values]

    next_value_at_risk_filtered = -next_cond_mean.values - np.sqrt(next_cond_var).values * next_quantiles.values
    next_value_at_risk_filtered = pd.DataFrame(
        next_value_at_risk_filtered, index=['1%', '5%'], columns=next_cond_var.columns) 
    
    next_value_at_risk = next_value_at_risk/100
    next_value_at_risk_filtered = next_value_at_risk_filtered/100
    mc_value_at_risk = np.abs(mc_value_at_risk)/100
    
    next_value_at_risk.to_excel(writer,'Var',startrow = 6,startcol=3,index=True,index_label='C.I.')
    next_value_at_risk_filtered.to_excel(writer,'Var',startrow = 15,startcol=3,index=True,index_label='C.I.')
    mc_value_at_risk.to_excel(writer,'Var',startrow = 24,startcol=3,index=True,index_label='C.I.')
#########################################################################
    #limits
#########################################################################    
    ta = pd.DataFrame(data_txt.sum(1),columns=['TOTAL ASSETS'])

    perc_ta = data_txt/np.tile(ta,np.size(data_txt,1))
    perc_ta.set_index([date_read],inplace=True)
    
    perc_ta_last = perc_ta.iloc[-1,:]
    
    bond_exp = perc_ta_last['BOND L'] + perc_ta_last['BOND NL']
    sub_exp =(( bond_ranking.loc['Subordinated'][0]+bond_ranking.loc['Jr Subordinated'][0])/ta.iloc[-1]).values[0]
    aim_exp = (( equity_market.loc['AIM ITALIA'][0])/ta.iloc[-1]).values[0]
    
    perc_ta_last = perc_ta_last.append(pd.DataFrame([bond_exp,sub_exp,aim_exp],index=['BOND','BOND SUB','AIM (ex WRT)']))
    perc_ta_last.columns = ['Exp %']
    
    
    idx_lim = ['1.Cash','   2.1.a Bond Listed','   2.1.b Bond Not Listed','3 Equity','4 Govi','5 ETF',\
               '6 Others','2 Bond','   2.2 Bond Subordinated','   3.1 *Equity AIM (ex wrt)']
    perc_ta_last.set_index([idx_lim],inplace=True)
    perc_ta_last['sort'] =[1,3,4,6,8,9,10,2,5,7]
    perc_ta_last['threshold'] = [0.1,np.nan,0.1,0.35,0.3,0.10,np.nan,0.6,0.10,0.05]
    perc_ta_last['+/-'] = ['<','<','<','<','<','<','<','<','<','>']
    
    perc_ta_last.sort_values(by='sort',inplace =True)
    
    ops = {'<':operator.lt,'>':operator.gt}
    for i in perc_ta_last.index:
        if np.isnan(perc_ta_last.loc[i,'threshold']):
            flag = 'nan'
        else:
            flag = ops[perc_ta_last.loc[i,'+/-']](perc_ta_last.loc[i,'Exp %'] ,perc_ta_last.loc[i,'threshold'])
        if flag=='nan':
            perc_ta_last.loc[i,'P/F'] = '-'
        elif flag:
            perc_ta_last.loc[i,'P/F'] = 'Passed'
        elif flag==False:
            perc_ta_last.loc[i,'P/F'] = 'Failed'
    
    perc_ta_last.drop(['sort'],axis=1,inplace=True)
    
    perc_ta_last.to_excel(writer,'Var',startrow = 32,startcol=2,index=True,index_label='Item',na_rep='-')

#########################################################################
    #MARKET SUMMARY
######################################################################### 

    y_n = dt.date.today().year
    y_p = y_n-1
    y_f = y_n + 1
    
    str_monitor = 'ISIN UNIVOCI.csv'
    
    mon = pd.read_csv(str_monitor,encoding = "ISO-8859-1",skiprows=2).dropna(how='all',axis=0)
    
    past_cap = ['CAP_5D','CAP_MTD','CAP_3M','CAP_6M','CAP_YTD','CAP_1YR']
    past_price = ['Price5D','PriceMTD','Price3M','Pice6M','PriceYTD','Price1YR']
    for i in range(len(past_cap)):
        mon[past_cap[i]] = mon['SHARES'] * mon[past_price[i]]
    
    margin = ['EBITDA_','NI_','FCF_']
    margin = [x+str(y) for x in margin for y in [y_p,y_n,y_f]]
    rev = ['REVENUE_'+str(y) for y in [y_p,y_n,y_f]]
    
    for i in range(len(margin)):
        for j in rev:
            if margin[i] not in mon.columns:
                pass
            elif j[-4:]==margin[i][-4:]:
                mon[margin[i]+'_margin'] =mon[margin[i]]/mon[j]
                
    mon['REV_GROWTH_' + str(y_n)] = mon['REVENUE_'+ str(y_n)]/ mon['REVENUE_'+ str(y_p)]-1
    mon['REV_GROWTH_' + str(y_f)] = mon['REVENUE_'+ str(y_f)]/mon['REVENUE_'+ str(y_n)]-1
    
    mon['EPS_GROWTH_' + str(y_n)] = mon['EPS_'+ str(y_n)]/ mon['EPS_'+ str(y_p)]-1
    mon['EPS_GROWTH_' + str(y_f)] = mon['EPS_'+ str(y_f)]/mon['EPS_'+ str(y_n)]-1
    
    index_mon = pd.unique(mon['INDEX']).tolist()
    index_mon.sort()
    industry_mon = pd.unique(mon['SETTORE INDUSTRIA']).tolist()
    
    index_YTD={}
    index_MTD={}
    index_weekly={}
    for i in index_mon:
        mon['IS_IN_' + i] = mon['INDEX']==i
        p_ytd = mon[mon['INDEX']==i][['PERF. YTD','TICKER','DESCRIZIONE']].sort_values(by='PERF. YTD',ascending=False)
        p_ytd.dropna(inplace=True)
        
        p_mtd = mon[mon['INDEX']==i][['PERF. MTD','TICKER','DESCRIZIONE']].sort_values(by='PERF. MTD',ascending=False)
        p_mtd.dropna(inplace=True)   
    
        p_5D = mon[mon['INDEX']==i][['PERF. 5D','TICKER','DESCRIZIONE']].sort_values(by='PERF. 5D',ascending=False)
        p_5D.dropna(inplace=True)   
        
        index_YTD['mon_'+i] = [p_ytd.iloc[:10,:],p_ytd.iloc[-10:,:].sort_values(by='PERF. YTD',ascending=True)]
        index_MTD['mon_'+i] = [p_mtd.iloc[:10,:],p_mtd.iloc[-10:,:].sort_values(by='PERF. MTD',ascending=True)]
        index_weekly['mon_'+i] = [p_5D.iloc[:10,:],p_5D.iloc[-10:,:].sort_values(by='PERF. 5D',ascending=True)]
        
    industry_YTD={}
    industry_MTD={}
    industry_weekly={}
    
    for i in industry_mon:
        i_ytd = mon[mon['SETTORE INDUSTRIA']==i][['PERF. YTD','TICKER','DESCRIZIONE','INDEX','MARKET CAP']].sort_values(by='PERF. YTD',ascending=False)
        i_ytd.dropna(inplace=True)
        
        i_mtd = mon[mon['SETTORE INDUSTRIA']==i][['PERF. MTD','TICKER','DESCRIZIONE','INDEX','MARKET CAP']].sort_values(by='PERF. MTD',ascending=False)
        i_mtd.dropna(inplace=True)   
    
        i_5D = mon[mon['SETTORE INDUSTRIA']==i][['PERF. 5D','TICKER','DESCRIZIONE','INDEX','MARKET CAP']].sort_values(by='PERF. 5D',ascending=False)
        i_5D.dropna(inplace=True)   
        
        industry_YTD['mon_'+i] = [i_ytd.iloc[:10,:],i_ytd.iloc[-10:,:].sort_values(by='PERF. YTD',ascending=True)]
        industry_MTD['mon_'+i] = [i_mtd.iloc[:10,:],i_mtd.iloc[-10:,:].sort_values(by='PERF. MTD',ascending=True)]
        industry_weekly['mon_'+i] = [i_5D.iloc[:10,:],i_5D.iloc[-10:,:].sort_values(by='PERF. 5D',ascending=True)]
    
    
    
    #industry_YTD['mon_'+i][1].to_excel(writer)
    #industry_YTD['mon_'+i][0].to_excel(writer,startcol=10,index_label=i)
    #
    
    perf_dict = [index_YTD,index_MTD,index_weekly]
    
    n_ind = 0
    for k in range(len(perf_dict)):
        col = 0
        for v in perf_dict[k].keys():
            
            for q in range(2):
                if q ==0:
                    ind_lab = v.replace('mon','BEST')
                else:
                    ind_lab = v.replace('mon','WORST')
                    
                n = len(perf_dict[k][v][q]) 
                            
                new_index = [x for x in range(1,n+1)]
                perf_dict[k][v][q].set_index([new_index],inplace=True)
                perf_dict[k][v][q].to_excel(writer,'Equity_market',startrow=n_ind,startcol=col,index_label=ind_lab)
                col+=5
        n_ind = n_ind +13
        
    ind_dict = [industry_YTD,industry_MTD,industry_weekly]
    
    n_ind = 0
    for k in range(len(ind_dict)):
        col = 0
        for v in ind_dict[k].keys():
            
            for q in range(2):
                if q ==0:
                    ind_lab = v.replace('mon','BEST')
                else:
                    ind_lab = v.replace('mon','WORST')
                    
                n = len(ind_dict[k][v][q]) 
                            
                new_index = [x for x in range(1,n+1)]
                ind_dict[k][v][q].set_index([new_index],inplace=True)
                ind_dict[k][v][q].to_excel(writer,'Equity_industry',startrow=n_ind,startcol=col,index_label=ind_lab)
                col+=7
        n_ind = n_ind +13 
        
    mon['PERF. YTD W'] = mon['PERF. YTD'] *  mon['CAP_YTD']/mon['CAP_YTD'].sum()
    mon['PERF. MTD W'] = mon['PERF. MTD'] *  mon['CAP_MTD']/mon['CAP_MTD'].sum()
    
    beta_industry_mean = mon.groupby(['SETTORE INDUSTRIA'])[['BETA']].mean()
    beta_industry_std = mon.groupby(['SETTORE INDUSTRIA'])[['BETA']].std()
    
    ind_stats = {'SETTORE INDUSTRIA':'count',
    #            'MARKET CAP':'sum',
                'MARKET CAP':'mean',
                 'PERF. YTD': 'mean',
                 'PERF. YTD W': 'sum',
                 'PERF. MTD': 'mean',
                 'PERF. MTD W':'sum',
                 'PERF. 5D': 'mean',
                 'BETA':'mean',
                 'PB_2018':'mean',
                 'IS_IN_AIM ITALIA':'sum',
                 'IS_IN_FTSE MIB':'sum',
                 'IS_IN_MID CAP':'sum',
                 'IS_IN_SMALL CAP':'sum',}
    
    sum_stats = mon.groupby(['SETTORE INDUSTRIA']).agg(ind_stats)
    sum_stats.to_excel(writer,'Performance sheet',startrow=35,startcol=2,index_label='Industry')

#########################################################################
    #CLOSE AND SAVE
#########################################################################   
    writer.save()

    book_template = xw.Book(name_to_save)
    perf_sheet = book_template.sheets['Perf Chart']
    
#########################################################################
    #plot performance  - YTD
#########################################################################
    fig,ax = plt.subplots(figsize=(20,13),dpi=dpi_report)
    date_YTD = dt.date(2018,12,28)
    x_axis = np.arange(len(bond_perf.set_index([timestamp_index_long]).loc[date_YTD:].index.tolist()))
    
    bond_YTD_plot = bond_perf.set_index([timestamp_index_long]).loc[date_YTD:].\
                divide(bond_perf.set_index([timestamp_index_long]).loc[date_YTD])
    equity_YTD_plot = eqt_perf.set_index([timestamp_index_long]).loc[date_YTD:].\
                divide(eqt_perf.set_index([timestamp_index_long]).loc[date_YTD])
    pir_YTD_plot = pir_perf_a.iloc[2:,:].set_index([timestamp_index_long]).loc[date_YTD:].\
                divide(pir_perf_a.iloc[2:,:].set_index([timestamp_index_long]).loc[date_YTD])
    
    ax.plot(x_axis,np.array(bond_YTD_plot),color='deepskyblue',lw=1.5)
    ax.plot(x_axis,np.array(equity_YTD_plot),color='navy',lw=1.5)
    ax.plot(x_axis,np.array(pir_YTD_plot),color='lightslategrey',lw=1)
    ax.set_title("Performance YTD",fontdict =  {'family': 'serif',
        'color':  'darkblue',
        'weight': 'normal',
        'size': 24,
        }) 
    ax.tick_params(axis = 'both', which = 'major', labelsize = 20)
    lg = ["BOND: %4.2f%%"%((bond_YTD_plot.iloc[-1][0]-1)*100),"EQUITY: %4.2f%%"%((equity_YTD_plot.iloc[-1][0]-1)*100),
          "PIR: %4.2f%%"%((pir_YTD_plot.iloc[-1][0]-1)*100)]
    plt.legend(lg,fontsize=20)
    equidate_ax(fig, ax, bond_perf.set_index([timestamp_index_long]).loc[date_YTD:].index)
#    plt.margins(1)
    
    plt.tight_layout(0)
    
    top = 0
    
    perf_sheet.pictures.add(fig,name='Performance YTD',top = top,update=True)
    top += fig.get_size_inches()[1]*100
    plt.close(fig)

#########################################################################
    #plot performance attribution - 2018 - YTD - MTD
#########################################################################
    w_YTD = w_asset_class.loc[int(date_YTD.strftime('%Y%m%d'))]
    ret_attribution = np.array([equity_YTD_plot.iloc[-1,0]-1,bond_YTD_plot.iloc[-1,0]-1])
    #YTD
    perf_attr =w_YTD * ret_attribution
    perf_attr = perf_attr.to_frame('YTD')
    perf_attr = perf_attr.rename({'eqt_cap':'EQT','bond_cap':'BOND'})*100
    #2018
#    perf_attr_2018 = np.array(w_asset_class[(w_asset_class.index>20171231)&(w_asset_class.index<20190101)].mean()) * np.array(perf_summary.loc[['Equity','Bond'],['2018']]).T
#    perf_attr_2018 = pd.DataFrame(perf_attr_2018.T,columns=['2018'],index=perf_attr.index)
    
    perf_attr_2018 = np.array(perf_summary.loc[['Equity','Bond','PIR'],['2018']]).T
    weq2018 = (perf_attr_2018[0,2] - perf_attr_2018[0,1])/(perf_attr_2018[0,0]-perf_attr_2018[0,1])
    wbond2018 = 1 - weq2018
    
    perf_attr_2018 = perf_attr_2018[0,:-1] * np.array([weq2018,wbond2018])
    perf_attr_2018 = pd.DataFrame(perf_attr_2018,columns=['2018'],index=perf_attr.index)
    
    #MTD
    perf_attr_MTD = np.array(perf_summary.loc[['Equity','Bond','PIR'],['MTD']]).T
    weqM = (perf_attr_MTD[0,2] - perf_attr_MTD[0,1])/(perf_attr_MTD[0,0]-perf_attr_MTD[0,1])
    wbondM = 1 - weqM
    
    perf_attr_MTD = perf_attr_MTD[0,:-1] * np.array([weqM,wbondM])
    perf_attr_MTD = pd.DataFrame(perf_attr_MTD,columns=['MTD'],index=perf_attr.index)
#    perf_attr = pd.concat([perf_attr_2018,perf_attr,perf_attr_MTD],axis=1)
    
    #2017
    perf_attr_2017 = np.array(perf_summary.loc[['Equity','Bond','PIR'],['2017']]).T
    weq2017 = (perf_attr_2017[0,2] - perf_attr_2017[0,1])/(perf_attr_2017[0,0]-perf_attr_2017[0,1])
    wbond2017 = 1 - weq2017
    
    perf_attr_2017 = perf_attr_2017[0,:-1] * np.array([weq2017,wbond2017])
    perf_attr_2017 = pd.DataFrame(perf_attr_2017,columns=['2017'],index=perf_attr.index)
    perf_attr = pd.concat([perf_attr_2017,perf_attr_2018,perf_attr,perf_attr_MTD],axis=1)
    
    fig1,ax1 = plt.subplots(figsize=(20,13),dpi=dpi_report)

    perf_attr.T.plot.bar(stacked=True,color=['navy','deepskyblue'],ax=ax1,width=0.4,legend=True)
    ax1.tick_params(axis = 'both', which = 'major', labelsize = 20)
    ax1.yaxis.set_major_formatter(mtick.PercentFormatter())
    ax1.set_title("Performance Attribution",fontdict =  {'family': 'serif',
        'color':  'darkblue',
        'weight': 'normal',
        'size': 24,
        }) 
#    ax1.set_ylim(-20,+20)
    plt.legend(fontsize=20)
    plt.tight_layout(0)
    perf_sheet.pictures.add(fig1,name='Performance attribution',top=top,update=True)
    top += fig1.get_size_inches()[1]*100
    plt.close(fig1)

#########################################################################
    #plot performance attribution - Monthly
#########################################################################
    
    df_monthly_perf_weights = []
    for i in range(np.size(df_monthly_perf,0)):
        weq = (df_monthly_perf.iloc[i,2] - df_monthly_perf.iloc[i,1])/(df_monthly_perf.iloc[i,0]-df_monthly_perf.iloc[i,1])
        wbond = 1-weq
        w_attr = [weq,wbond]
        df_monthly_perf_weights.append(w_attr)
        
    df_monthly_perf_weights1 = pd.DataFrame(df_monthly_perf_weights,columns=['EQT','BOND'],index=df_monthly_perf.index)
    
    perf_attr_monthly = np.array(df_monthly_perf.iloc[:,:2]) * np.array(df_monthly_perf_weights1)
    perf_attr_monthly = pd.DataFrame(perf_attr_monthly,columns=['EQT','BOND'],index=df_monthly_perf.index )*100
    
    fig2,ax2 = plt.subplots(figsize=(20,13),dpi=dpi_report)
    
    perf_attr_monthly.plot.bar(stacked=True,color=['navy','deepskyblue'],ax=ax2,width=0.4,legend=True)
    
    ax2.tick_params(axis = 'both', which = 'major', labelsize = 20)
    ax2.yaxis.set_major_formatter(mtick.PercentFormatter())
    ax2.set_title("Performance Attribution Monthly",fontdict =  {'family': 'serif',
        'color':  'darkblue',
        'weight': 'normal',
        'size': 24,
        }) 
    plt.tight_layout(0)
    plt.legend(fontsize=20)
    perf_sheet.pictures.add(fig2,name='Performance attribution monthly',top=top,update=True)
    top +=fig2.get_size_inches()[1]*100
    plt.close(fig2)

#########################################################################
    #plot performance overall
#########################################################################
    
    fig3,ax3 = plt.subplots(figsize=(20,13),dpi=dpi_report)
    date_inc = dt.date(2017,7,7)
    x_axis = np.arange(len(bond_perf.set_index([timestamp_index_long]).loc[date_inc:].index.tolist()))
    
    bond_inc_plot = bond_perf.set_index([timestamp_index_long]).loc[date_inc:].\
                divide(bond_perf.set_index([timestamp_index_long]).loc[date_inc])
    equity_inc_plot = eqt_perf.set_index([timestamp_index_long]).loc[date_inc:].\
                divide(eqt_perf.set_index([timestamp_index_long]).loc[date_inc])
    pir_inc_plot = pir_perf_a.iloc[2:,:].set_index([timestamp_index_long]).loc[date_inc:].\
                divide(pir_perf_a.iloc[2:,:].set_index([timestamp_index_long]).loc[date_inc])
    
    ax3.plot(x_axis,np.array(bond_inc_plot),color='deepskyblue',lw=1.5)
    ax3.plot(x_axis,np.array(equity_inc_plot),color='navy',lw=1.5)
    ax3.plot(x_axis,np.array(pir_inc_plot),color='lightslategrey',lw=1)
    ax3.set_title("Performance since Inception",fontdict =  {'family': 'serif',
        'color':  'darkblue',
        'weight': 'normal',
        'size': 24,
        }) 
    ax3.tick_params(axis = 'both', which = 'major', labelsize = 20)
    lg = ["BOND: %4.2f%%"%((bond_inc_plot.iloc[-1][0]-1)*100),"EQUITY: %4.2f%%"%((equity_inc_plot.iloc[-1][0]-1)*100),
          "PIR: %4.2f%%"%((pir_inc_plot.iloc[-1][0]-1)*100)]
    plt.legend(lg,fontsize=20)
    equidate_ax(fig3, ax3, bond_perf.set_index([timestamp_index_long]).loc[date_inc:].index)   
    plt.tight_layout(0)
    perf_sheet.pictures.add(fig3,name='Performance inception',top=top,update=True)
    top += fig3.get_size_inches()[1]*100
    plt.close(fig3)

    

    
#    p2_a = beuty_pie(fund_industry,"Fund - Industry Breakdown")
#    p3_a = beuty_pie(fund_ac,"Fund - Asset Class")
#    p4_a = beuty_pie(fund_listed,"Fund -Listed")
#    p5_a = beuty_pie(bond_industry,"Bond - Industry Breakdown")
#    p6_a = beuty_pie(bond_maturity,"Bond - Maturity")
#    p7_a = beuty_pie(bond_duration,"Bond - Duration")
#    p8_a = beuty_pie(bond_ranking,"Bond - Security Type")
#    p9_a = beuty_pie(bond_coupon,"Bond - Coupon Type")
#    p10_a = beuty_pie(equity_industry,"Equity - Industry Breakdown")
#    p11_a = beuty_pie(equity_ranking,"Equity - Security Type")
#    p12_a = beuty_pie(equity_market,"Equity - Index")
#    p13_a = beuty_pie(equity_dividend,"Equity - Dividend")
#    p14_a = beuty_pie(equity_cap,"Equity - Capitalisation")    
#    
#    p_list = [p2_a,p3_a,p4_a,p5_a,p6_a,p7_a,p8_a,p9_a,p10_a,p11_a,p12_a,p13_a,p14_a,]
#    
#    for pic in p_list:
#        perf_sheet.pictures.add(pic[0],name = pic[1],top=top,update=True)
#        top += pic[0].get_size_inches()[1]*100/2
    
    title_equity_pic =  ["Equity - Industry Breakdown","Equity - Security Type","Equity - Index","Equity - Capitalisation"]
    title_bond_pic =  ["Bond - Industry Breakdown","Bond - Security Type","Bond - Coupon Type","Bond - Maturity"]
    title_fund_pic = ["Fund - Industry Breakdown","Fund - Asset Class"]
    
    equity_pic_report = beauty_pie_4(equity_industry,equity_ranking,equity_market,equity_cap,title_equity_pic)
    bond_pic_report =  beauty_pie_4(bond_industry,bond_ranking,bond_coupon,bond_maturity,title_bond_pic)
    fund_pic_report = beauty_pie_2(fund_industry,fund_ac,title_fund_pic)
    
    perf_sheet.pictures.add(equity_pic_report,name = 'EQUITY PIC REPORT',top=top,update=True,save_with_document=True)
    top += equity_pic_report.get_size_inches()[1]*100
    
    perf_sheet.pictures.add(bond_pic_report,name = 'BOND PIC REPORT',top=top,update=True,save_with_document=True)
    top += bond_pic_report.get_size_inches()[1]*100
    
    perf_sheet.pictures.add(fund_pic_report,name = 'FUND PIC REPORT',top=top,update=True,save_with_document=True)
    top += fund_pic_report.get_size_inches()[1]*100 

    width =1.1
    rot=0
    
    
#########################################################################
    #plot SEASONALITY
#########################################################################
    fig, axes = plt.subplots(figsize=(25,15),nrows=3, ncols=1,dpi=dpi_report)
    
    df_fund_perf.T.drop('YTD').plot.bar(width=width,rot=rot,ax=axes[0])
    df_equity_perf.T.drop('YTD').plot.bar(width=width,rot=rot,ax=axes[1])
    df_bond_perf.T.drop('YTD').plot.bar(width=width,rot=rot,ax=axes[2])
    
    plt.setp([x.get_xticklabels() for x in axes] ,size=20,  color='navy')
    plt.setp([x.get_yticklabels() for x in axes] ,size=15, color='k')
    
    vals = [x.get_yticks() for x in axes]
    [x.yaxis.set_major_formatter(mtick.PercentFormatter(1.0))for x in axes]
    [x.legend(loc=8,ncol=len_index,fontsize=15)for x in axes]
    
    axes[0].set_ylabel("FUND",fontsize=18,color='navy')
    axes[1].set_ylabel("EQUITY",fontsize=18,color='navy')
    axes[2].set_ylabel("BOND",fontsize=18,color='navy')

    perf_sheet.pictures.add(fig,name = 'SEASONALITY',top=top,update=True,save_with_document=True)
    top += fig.get_size_inches()[1]*100 

#########################################################################
    #plot TOP 10 HOLDINGs
#########################################################################
    df_top_hold =(pir_portfolio.sort_values(['MKT VALUE'], ascending=False)).iloc[:10,:]
    df_top_hold['MKT VALUE']=df_top_hold['MKT VALUE'].divide(pir_portfolio['MKT VALUE'].sum())*100

    fig2,ax2 = plt.subplots(num=None,figsize=(20,13),dpi=dpi_report)
    df_top_hold.set_index('SECURITY NAME')['MKT VALUE'].plot.barh( color='navy',ax=ax2)
    fmt = '%.1f%%' # Format you want the ticks, e.g. '40%'
    ax2.set_title("TOP 10 HOLDINGS ",fontdict =  {'family': 'serif',
        'color':  'darkblue',
        'weight': 'bold',
        'size': 24,
        }) 
    ax2.yaxis.label.set_visible(False)
    plt.gca().invert_yaxis()
    xticks = mtick.FormatStrFormatter(fmt)
    ax2.xaxis.set_major_formatter(xticks)
    plt.tight_layout()
    
    for i in range(10):
        plt.text(df_top_hold.iloc[i,8]-0.35,i+0.05,('%.2f%%')%df_top_hold.iloc[i,8],fontdict =  {'family': 'arial',
        'color':  'white',
        'weight': 'bold',
        'size': 20,
        }) 
    plt.setp(ax2.get_yticklabels(),size=20,  color='navy')
    plt.setp(ax2.get_xticklabels(),size=20,  color='navy')
    
    perf_sheet.pictures.add(fig2,name = 'TOP 10',top=top,update=True,save_with_document=True)
    top += fig.get_size_inches()[1]*100 
    plt.close(fig2)
#########################################################################
    #plot TOP 10 ISSUERS
#########################################################################
    df_top_issuer =(pir_portfolio.sort_values(['Esp. Eqt+Bond'], ascending=False)).drop_duplicates(['Esp. Eqt+Bond']).iloc[:10,:]
    fig3,ax3 = plt.subplots(num=None,figsize=(20,13),dpi=dpi_report)
    (df_top_issuer.set_index('ISSUER')['Esp. Eqt+Bond']*100).plot.barh( color='navy',ax=ax3)
    fmt = '%.1f%%' # Format you want the ticks, e.g. '40%'
    ax3.set_title("TOP 10 ISSUERS ",fontdict =  {'family': 'serif',
        'color':  'darkblue',
        'weight': 'bold',
        'size': 24,
        }) 
    ax3.yaxis.label.set_visible(False)
    plt.gca().invert_yaxis()
    xticks = mtick.FormatStrFormatter(fmt)
    ax3.xaxis.set_major_formatter(xticks)
    plt.tight_layout()
    
    for i in range(10):
        plt.text(df_top_issuer.iloc[i,9]*100-0.45,i+0.05,('%.2f%%')%(df_top_issuer.iloc[i,9]*100),fontdict =  {'family': 'arial',
        'color':  'white',
        'weight': 'bold',
        'size': 20,
        }) 
    plt.setp(ax3.get_yticklabels(),size=20,  color='navy')
    plt.setp(ax3.get_xticklabels(),size=20,  color='navy')
    
    perf_sheet.pictures.add(fig3,name = 'TOP 10 ISSUERS',top=top,update=True,save_with_document=True)
    top += fig3.get_size_inches()[1]*100  
    plt.close(fig3)

#########################################################################
    #STACKED CHART
#########################################################################

    
    fig,ax = plt.subplots(figsize=(20,13),dpi=dpi_report)
    ax.set_title("% Fund Allocation",fontdict =  {'family': 'serif',
        'color':  'darkblue',
        'weight': 'bold',
        'size': 24,
        }) 
    
    
    perc.plot(kind='bar',stacked=True,ax=ax,xticks=None)
    pos = np.arange(len(perc.index))
    ticks = plt.xticks(pos[::20], date[::20], rotation=90)
    plt.ylim([0,1])
    
    plt.setp(ax.get_yticklabels(),size=20,  color='navy')
    plt.setp(ax.get_xticklabels(),size=20,  color='navy')
    
    vals = ax.get_yticks()
    ax.set_yticklabels(['{:,.0%}'.format(x) for x in vals])
    plt.legend(fontsize=20)

    plt.tight_layout()
    
    perf_sheet.pictures.add(fig,name = 'FUND ALLOCATION TS',top=top,update=True,save_with_document=True)
    top += fig3.get_size_inches()[1]*100     
    plt.close(fig)

    
    
#########################################################################
    #comparison vs last year
#########################################################################
    idx_ytd_txt = date_fl.get_loc(date_YTD)
    width = 0.35
    
    fig,ax = plt.subplots(figsize=(20,13),dpi=dpi_report)
    
    r1 =  np.arange(np.size(data_txt,1))
    r2 = [x + width for x in r1]
    
    ax.bar(r1,data_txt.iloc[idx_ytd_txt,:],width=width,)
    ax.bar(r2,data_txt.iloc[-1,:],width=width,color='navy')
    
    plt.xticks([(r + width/2) for r in range(np.size(data_txt,1))],[x for x  in data_txt.columns])
    
    ylabels = ['{:,.0f}'.format(x) + 'K' for x in ax.get_yticks()/1000]
    ax.set_yticklabels(ylabels)
    ax.yaxis.label.set_visible(False)
    
    plt.setp(ax.get_yticklabels(),size=20,  color='navy')
    plt.setp(ax.get_xticklabels(),size=20,  color='navy')
    
    ax.set_title("ASSETS: Last Year vs Now ",fontdict =  {'family': 'serif',
        'color':  'darkblue',
        'weight': 'bold',
        'size': 24,
        }) 
    
    plt.legend([ date_fl[idx_ytd_txt].strftime('%d-%m-%y'), date_fl[-1].strftime('%d-%m-%y')],fontsize=20)
    
    perf_sheet.pictures.add(fig,name = 'LAST YEAR ',top=top,update=True,save_with_document=True)
    top += fig3.get_size_inches()[1]*100     
    plt.close(fig)


    n_d = 30
#============================================================
    #rolling corr
#============================================================

    X_roll = berk_ret['FTSEMIB Index']
    X_roll_reg = np.vstack((np.ones_like(X_roll),X_roll)).T
    
    roll_corr = np.zeros((np.size(ret_all,0)-n_d,3))
    se_corr = np.zeros((np.size(ret_all,0)-n_d,3))
    roll_beta = np.zeros((np.size(ret_all,0)-n_d,3))
    se_beta = np.zeros((np.size(ret_all,0)-n_d,3))
    for i in range(np.size(ret_all,0)-n_d):
        bench_slice = X_roll.iloc[i:i+n_d]
        X_reg = X_roll_reg[i:i+n_d]
        ret_slice = ret_all.iloc[i:i+n_d,:]
        for j in range(3):
            coeff = np.linalg.inv(X_reg.T @ X_reg) @ (X_reg.T @ np.array(ret_slice)[:,j])
            beta_roll = coeff[1]  
            corr_i = (np.corrcoef(bench_slice,ret_slice.iloc[:,j]))[1,0]
            
            se = ((1-corr_i**2)/(n_d-2))**0.5
            var_hat = ((ret_slice.iloc[:,j] - X_reg@coeff)**2).sum()
            se_reg =((var_hat/(n_d-2))/(((X_reg[:,1]-np.mean(X_reg[:,1]))**2).sum()))**0.5
            
            roll_corr[i,j] = corr_i
            se_corr[i,j] = se
            roll_beta[i,j] = beta_roll
            se_beta[i,j] = se_reg
            
            
    date_index_roll = X_roll.index[n_d:].tolist()
    date_index_roll = [dt.datetime.strptime(str(x),'%Y%m%d') for x in date_index_roll]
    
    roll_corr = pd.DataFrame(roll_corr,index=date_index_roll,columns = ['EQUITY','BOND','FUND'])
    roll_se_p = roll_corr + se_corr
    roll_se_m = roll_corr - se_corr
    
    fig,ax = plt.subplots(ncols=1,nrows=3,figsize=(16,9),sharex=True,dpi=dpi_report)
    
    roll_corr['FUND'].plot(x=date_index_roll,ax=ax[0],color='navy')
    roll_corr['EQUITY'].plot(x=date_index_roll,ax=ax[1],color='navy')
    roll_corr['BOND'].plot(x=date_index_roll,ax=ax[2],color='navy')
    
    
    plt.setp([x.get_xticklabels() for x in ax] ,size=20,  color='navy')
    plt.setp([x.get_yticklabels() for x in ax] ,size=15, color='k')
      
    ax[0].set_ylabel("FUND",fontsize=18,color='navy')
    ax[1].set_ylabel("EQUITY",fontsize=18,color='navy')
    ax[2].set_ylabel("BOND",fontsize=18,color='navy')
    
    ax[0].fill_between(roll_corr.index,roll_se_p['FUND'],roll_se_m['FUND'],color='khaki')
    ax[1].fill_between(roll_corr.index,roll_se_p['EQUITY'],roll_se_m['EQUITY'],color='khaki')
    ax[2].fill_between(roll_corr.index,roll_se_p['BOND'],roll_se_m['BOND'],color='khaki')
    
    
    ax[0].set_title(str(n_d) +"-day Rolling Correlation vs FTSEMIB",fontdict =  {'family': 'serif',
        'color':  'darkblue',
        'weight': 'bold',
        'size': 24,
        }) 
    
    plt.tight_layout()   
    
    perf_sheet.pictures.add(fig,name = 'ROLLING CORR',top=top,update=True,save_with_document=True)
    top += fig.get_size_inches()[1]*100     
    plt.close(fig)

    
    roll_beta = pd.DataFrame(roll_beta,index=date_index_roll,columns = ['EQUITY','BOND','FUND'])
    rollb_se_p = roll_beta + se_beta
    rollb_se_m = roll_beta - se_beta
    
    fig,ax = plt.subplots(ncols=1,nrows=3,figsize=(16,9),sharex=True,dpi=dpi_report)
    
    roll_beta['FUND'].plot(x=date_index_roll,ax=ax[0],color='r')
    roll_beta['EQUITY'].plot(x=date_index_roll,ax=ax[1],color='r')
    roll_beta['BOND'].plot(x=date_index_roll,ax=ax[2],color='r')
    
    
    plt.setp([x.get_xticklabels() for x in ax] ,size=20,  color='navy')
    plt.setp([x.get_yticklabels() for x in ax] ,size=15, color='k')
      
    ax[0].set_ylabel("FUND",fontsize=18,color='navy')
    ax[1].set_ylabel("EQUITY",fontsize=18,color='navy')
    ax[2].set_ylabel("BOND",fontsize=18,color='navy')
    
    ax[0].fill_between(roll_beta.index,rollb_se_p['FUND'],rollb_se_m['FUND'],color='khaki')
    ax[1].fill_between(roll_beta.index,rollb_se_p['EQUITY'],rollb_se_m['EQUITY'],color='khaki')
    ax[2].fill_between(roll_beta.index,rollb_se_p['BOND'],rollb_se_m['BOND'],color='khaki')
    
    
    ax[0].set_title(str(n_d) +"-day Rolling BETA vs FTSEMIB",fontdict =  {'family': 'serif',
        'color':  'darkblue',
        'weight': 'bold',
        'size': 24,
        }) 
    
    plt.tight_layout()    
    
    perf_sheet.pictures.add(fig,name = 'ROLLING BETA',top=top,update=True,save_with_document=True)
    top += fig.get_size_inches()[1]*100     
    plt.close(fig)
    #==========================================
    #net subs by month
    #==========================================
    fig, ax = plt.subplots(figsize=(16,9),dpi=dpi_report)
    
    #ax = plt.bar(net_subs.index,net_subs.iloc[:,0])
    net_subs.plot.bar(ax=ax,color='tomato',legend=False)
    plt.setp(ax.get_xticklabels() ,size=20,  color='navy')
    plt.setp(ax.get_yticklabels() ,size=15, color='k')
    ax.yaxis.set_major_formatter(StrMethodFormatter('{x:,.0f}'))
    ax.set_title("MONTHLY SUBSCRIPTIONS (ex seed money)\n",fontdict =  {'family': 'serif',
        'color':  'darkblue',
        'weight': 'bold',
        'size': 24,
        }) 
    
    plt.tight_layout()
    perf_sheet.pictures.add(fig,name = 'NET SUBS',top=top,update=True,save_with_document=True)
    plt.close(fig)
    #==========================================
    #cumulative  subs by month
    #==========================================
    cum_subs = net_subs.cumsum()
    
    fig, ax = plt.subplots(figsize=(16,9),dpi=dpi_report)
    
    cum_subs.plot.bar(ax=ax,color='tomato',legend=False)
    
    plt.setp(ax.get_xticklabels() ,size=20,  color='navy')
    plt.setp(ax.get_yticklabels() ,size=15, color='k')
    ax.yaxis.set_major_formatter(StrMethodFormatter('{x:,.0f}'))
    ax.set_title("CUMULATIVE NET SUBSCRIPTIONS (ex seed money)\n",fontdict =  {'family': 'serif',
        'color':  'darkblue',
        'weight': 'bold',
        'size': 24,
        }) 
    plt.tight_layout()
    perf_sheet.pictures.add(fig,name = 'CUM NET SUBS',top=top,update=True,save_with_document=True)
    plt.close(fig)
      
#=====================================================  
    #PLOT VAR 1
#=====================================================
    
    fig,ax=plt.subplots(figsize=(16,10),dpi=dpi_report)
    
    ax = value_at_risk.plot(legend=False,ax=ax)
    xl = ax.set_xlim(value_at_risk.index[0], value_at_risk.index[-1])
    
    rets_split = pir[split_date:].copy()
    rets_split.name = 'PIR'
    
    c = []
    for idx in value_at_risk.index:
       
        if rets_split.loc[idx].values > -value_at_risk.loc[idx, '5%']:
            c.append('#000000')
        elif rets_split.loc[idx].values < -value_at_risk.loc[idx, '1%']:
            c.append('#BB0000')
        else:
            c.append('#BB00BB')
    
    c = np.array(c, dtype='object')
    labels = {
        '#BB0000': '1% Exceedence',
        '#BB00BB': '5% Exceedence',
        '#000000': 'No Exceedence'
    }
    markers = {'#BB0000': 'x', '#BB00BB': 's', '#000000': 'o'}
    for color in np.unique(c):
        sel = c == color
        ax.scatter(
            rets_split.index[sel],
            -rets_split.loc[sel],
            marker=markers[color],
            c=c[sel],
            label=labels[color])
    ax.set_title('Parametric VaR - GARCH(1,1), dist='+dist,fontdict =  {'family': 'serif',
            'color':  'darkblue',
            'weight': 'bold',
            'size': 24,
            }) 
    plt.tight_layout()
    leg = ax.legend(frameon=False, ncol=3,fontsize=15)
    plt.setp(ax.get_xticklabels() ,size=20,  color='navy')
    plt.setp(ax.get_yticklabels() ,size=15, color='k')
    perf_sheet.pictures.add(fig,name = 'VAR PARAMETRIC',top=top,update=True,save_with_document=True)
    plt.close(fig)
#=====================================================  
    #PLOT VAR FILTERED
#=====================================================
    fig,ax=plt.subplots(figsize=(16,10),dpi=dpi_report)
    
    ax = value_at_risk_empirical.plot(legend=False,ax=ax)
    xl = ax.set_xlim(value_at_risk_empirical.index[0], value_at_risk_empirical.index[-1])
    
    rets_split = pir[split_date:].copy()
    rets_split.name = 'PIR'
    
    c = []
    for idx in value_at_risk.index:
       
        if rets_split.loc[idx].values > -value_at_risk_empirical.loc[idx, '5%']:
            c.append('#000000')
        elif rets_split.loc[idx].values < -value_at_risk_empirical.loc[idx, '1%']:
            c.append('#BB0000')
        else:
            c.append('#BB00BB')
    
    c = np.array(c, dtype='object')
    labels = {
        '#BB0000': '1% Exceedence',
        '#BB00BB': '5% Exceedence',
        '#000000': 'No Exceedence'
    }
    markers = {'#BB0000': 'x', '#BB00BB': 's', '#000000': 'o'}
    for color in np.unique(c):
        sel = c == color
        ax.scatter(
            rets_split.index[sel],
            -rets_split.loc[sel],
            marker=markers[color],
            c=c[sel],
            label=labels[color])
    ax.set_title('Filtered Historical Simulation VaR - GARCH(1,1), dist=empirical',fontdict =  {'family': 'serif',
            'color':  'darkblue',
            'weight': 'bold',
            'size': 24,
            }) 
    plt.tight_layout()
    leg = ax.legend(frameon=False, ncol=3,fontsize=15)
    plt.setp(ax.get_xticklabels() ,size=20,  color='navy')
    plt.setp(ax.get_yticklabels() ,size=15, color='k')
    perf_sheet.pictures.add(fig,name = 'VAR FILTERED',top=top,update=True,save_with_document=True)
    
    
    book_template.save(name_to_save)
    book_template.save()
    book_template.close()    
    
if to_text == "y":
    sys.stdout.close()
    

    
sys.stdout = old_stdout
    


    
    
    

